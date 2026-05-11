from openpyxl import Workbook

from backend.app.services.outbound_reconciliation import (
    OutboundItem,
    _load_cutting_sheet,
    _load_shipping_sheet,
    load_outbound_items,
    outbound_summary,
    parse_outbound_text,
    query_outbound,
    reconcile_outbound_items,
)


def test_parse_outbound_text_extracts_order_part_and_quantity() -> None:
    items = parse_outbound_text("SO202604210135 | 6 | C.P.XS.000122001 | RTK", "shipping")

    assert len(items) == 1
    assert items[0].order_no == "SO202604210135"
    assert items[0].part_code == "C.P.XS.000122001"
    assert items[0].quantity == 6


def test_reconcile_outbound_items_compares_quantities() -> None:
    rows = reconcile_outbound_items(
        [OutboundItem("SO202604210135", "C.P.XS.000122001", 6, "cutting", "")],
        [OutboundItem("SO202604210135", "C.P.XS.000122001", 8, "shipping", "")],
    )

    assert rows[0].status == "over_shipped"
    assert rows[0].shipping_qty == 8


def test_load_workbook_business_sheets(tmp_path) -> None:
    path = tmp_path / "outbound.xlsx"
    workbook = Workbook()
    shipping = workbook.active
    shipping.title = "21单830个物料发货单-多物料订单"
    shipping.append(["出库单号", "数量", "备件编码", "备件名称"])
    shipping.append(["SO202604210135", 6, "C.P.XS.000122001", "RTK"])
    shipping.append([None, 2, "C.P.XS.000143004", "RTK-2"])
    cutting = workbook.create_sheet("拣货单-多物料订单")
    cutting.append(["数量", "备件编码", "备件名称", "库位"])
    cutting.append([6, "C.P.XS.000122001", "RTK", "A-B03-011"])
    workbook.save(path)

    shipping_items = _load_shipping_sheet(path, "21单830个物料发货单-多物料订单")
    cutting_items = _load_cutting_sheet(path, "拣货单-多物料订单")

    assert shipping_items[0].order_no == "SO202604210135"
    assert shipping_items[0].quantity == 6
    assert shipping_items[1].order_no == "SO202604210135"
    assert shipping_items[1].part_code == "C.P.XS.000143004"
    assert cutting_items[0].order_no == "PICKING_TOTAL"
    assert cutting_items[0].part_code == "C.P.XS.000122001"


def test_load_ordered_picking_sheet_preserves_orders_and_locations(tmp_path) -> None:
    path = tmp_path / "outbound.xlsx"
    workbook = Workbook()
    shipping = workbook.active
    shipping.title = "shipping"
    shipping.append(["出库单号", "数量", "备件编码", "备件名称"])
    shipping.append(["SO202605990001", 1, "C.P.XS.000999001", "Other"])
    cutting = workbook.create_sheet("96单小单")
    cutting.append(["出库单号", "数量", "备件编码", "备件名称", "库位", "", ""])
    cutting.append(["SO202605060078", 1, "W.E.CH.000028000", "YUKA 彩盒", "A-A06-013", "", ""])
    cutting.append(["SO202605030090", 1, "C.P.SH.000256000", "HM434 安全钥匙", "#VALUE!", "", ""])
    workbook.save(path)

    cutting_items = _load_cutting_sheet(path, "96单小单")

    assert cutting_items[0].order_no == "SO202605060078"
    assert cutting_items[0].locations == ("A-A06-013",)
    assert cutting_items[1].order_no == "SO202605030090"
    assert cutting_items[1].locations == ()


def test_ordered_picking_sheet_becomes_order_source_when_shipping_sheet_mismatches(
    tmp_path, monkeypatch
) -> None:
    path = tmp_path / "outbound.xlsx"
    workbook = Workbook()
    shipping = workbook.active
    shipping.title = "shipping"
    shipping.append(["出库单号", "数量", "备件编码", "备件名称"])
    shipping.append(["SO202605990001", 1, "C.P.XS.000999001", "Other"])
    cutting = workbook.create_sheet("96单小单")
    cutting.append(["出库单号", "数量", "备件编码", "备件名称", "库位"])
    cutting.append(["SO202605060078", 1, "W.E.CH.000028000", "YUKA 彩盒", "A-A06-013"])
    workbook.save(path)

    from backend.app.services import outbound_reconciliation

    class FakeSettings:
        outbound_workbook_file = path
        outbound_shipping_sheet = "shipping"
        outbound_cutting_sheet = "96单小单"
        outbound_cutting_text_file = tmp_path / "missing-cut.txt"
        outbound_shipping_text_file = tmp_path / "missing-ship.txt"

    monkeypatch.setattr(outbound_reconciliation, "get_settings", lambda: FakeSettings())

    _, shipping_items = load_outbound_items()
    summary = outbound_summary()

    assert [item.order_no for item in shipping_items] == ["SO202605060078"]
    assert summary["order_numbers"]["shipping"] == ["SO202605060078"]


def test_outbound_summary_uses_part_totals(tmp_path, monkeypatch) -> None:
    path = tmp_path / "outbound.xlsx"
    workbook = Workbook()
    shipping = workbook.active
    shipping.title = "ship"
    shipping.append(["出库单号", "数量", "备件编码", "备件名称"])
    shipping.append(["SO202604210135", 8, "C.P.XS.000122001", "RTK"])
    cutting = workbook.create_sheet("cut")
    cutting.append(["数量", "备件编码", "备件名称", "库位"])
    cutting.append([6, "C.P.XS.000122001", "RTK", "A-B03-011"])
    workbook.save(path)

    from backend.app.services import outbound_reconciliation

    class FakeSettings:
        outbound_workbook_file = path
        outbound_shipping_sheet = "ship"
        outbound_cutting_sheet = "cut"
        outbound_cutting_text_file = tmp_path / "missing-cut.txt"
        outbound_shipping_text_file = tmp_path / "missing-ship.txt"

    monkeypatch.setattr(outbound_reconciliation, "get_settings", lambda: FakeSettings())

    summary = outbound_summary()

    assert summary["part_rows"][0]["status"] == "over_shipped"
    assert summary["part_rows"][0]["difference"] == 2
    assert "rows" not in summary


def test_outbound_summary_marks_unreadable_quantities(tmp_path, monkeypatch) -> None:
    path = tmp_path / "outbound.xlsx"
    workbook = Workbook()
    shipping = workbook.active
    shipping.title = "ship"
    shipping.append(["出库单号", "数量", "备件编码", "备件名称"])
    shipping.append(["SO202604210135", "数量六", "C.P.XS.000122001", "RTK"])
    cutting = workbook.create_sheet("cut")
    cutting.append(["数量", "备件编码", "备件名称", "库位"])
    cutting.append([6, "C.P.XS.000122001", "RTK", "A-B03-011"])
    workbook.save(path)

    from backend.app.services import outbound_reconciliation

    class FakeSettings:
        outbound_workbook_file = path
        outbound_shipping_sheet = "ship"
        outbound_cutting_sheet = "cut"
        outbound_cutting_text_file = tmp_path / "stale-cut.txt"
        outbound_shipping_text_file = tmp_path / "stale-ship.txt"

    monkeypatch.setattr(outbound_reconciliation, "get_settings", lambda: FakeSettings())

    summary = outbound_summary()

    assert summary["part_rows"][0]["status"] == "quantity_unreadable"


def test_load_outbound_items_fails_when_workbook_missing(tmp_path, monkeypatch) -> None:
    from backend.app.services import outbound_reconciliation

    class FakeSettings:
        outbound_workbook_file = tmp_path / "missing.xlsx"
        outbound_shipping_sheet = "ship"
        outbound_cutting_sheet = "cut"
        outbound_cutting_text_file = tmp_path / "missing-cut.txt"
        outbound_shipping_text_file = tmp_path / "missing-ship.txt"

    monkeypatch.setattr(outbound_reconciliation, "get_settings", lambda: FakeSettings())

    try:
        load_outbound_items()
    except RuntimeError as exc:
        assert "outbound workbook not found" in str(exc)
    else:
        raise AssertionError("expected missing workbook to fail closed")


def test_load_outbound_items_ignores_stale_text_when_workbook_missing(
    tmp_path, monkeypatch
) -> None:
    cut_text = tmp_path / "cut.txt"
    ship_text = tmp_path / "ship.txt"
    cut_text.write_text("SO202604210135 1 C.P.XS.000122001", encoding="utf-8")
    ship_text.write_text("SO202604210135 1 C.P.XS.000122001", encoding="utf-8")

    from backend.app.services import outbound_reconciliation

    class FakeSettings:
        outbound_workbook_file = tmp_path / "missing.xlsx"
        outbound_shipping_sheet = "ship"
        outbound_cutting_sheet = "cut"
        outbound_cutting_text_file = cut_text
        outbound_shipping_text_file = ship_text

    monkeypatch.setattr(outbound_reconciliation, "get_settings", lambda: FakeSettings())

    try:
        load_outbound_items()
    except RuntimeError as exc:
        assert "outbound workbook not found" in str(exc)
    else:
        raise AssertionError("expected missing workbook to ignore stale OCR text")


def test_query_outbound_reports_selected_order_membership(tmp_path, monkeypatch) -> None:
    path = tmp_path / "outbound.xlsx"
    workbook = Workbook()
    shipping = workbook.active
    shipping.title = "ship"
    shipping.append(["出库单号", "数量", "备件编码", "备件名称"])
    shipping.append(["SO202604210135", 6, "C.P.XS.000122001", "RTK"])
    cutting = workbook.create_sheet("cut")
    cutting.append(["数量", "备件编码", "备件名称", "库位"])
    cutting.append([6, "C.P.XS.000122001", "RTK", "A-B03-011"])
    workbook.save(path)

    from backend.app.services import outbound_reconciliation

    class FakeSettings:
        outbound_workbook_file = path
        outbound_shipping_sheet = "ship"
        outbound_cutting_sheet = "cut"
        outbound_cutting_text_file = tmp_path / "missing-cut.txt"
        outbound_shipping_text_file = tmp_path / "missing-ship.txt"

    monkeypatch.setattr(outbound_reconciliation, "get_settings", lambda: FakeSettings())

    hit = query_outbound("C.P.XS.000122001", selected_orders=["SO202604210135"])
    miss = query_outbound("C.P.XS.000122001", selected_orders=["SO202604210999"])

    assert hit["belongs_to_selected"] is True
    assert hit["matching_selected_orders"][0]["order_no"] == "SO202604210135"
    assert miss["belongs_to_selected"] is False
    assert miss["matching_other_orders"][0]["order_no"] == "SO202604210135"


def test_query_outbound_expands_material_mapping_matches(tmp_path, monkeypatch) -> None:
    path = tmp_path / "outbound.xlsx"
    workbook = Workbook()
    shipping = workbook.active
    shipping.title = "ship"
    shipping.append(["出库单号", "数量", "备件编码", "备件名称"])
    shipping.append(["SO202604210135", 6, "C.G.LM.000011000", "Robot"])
    cutting = workbook.create_sheet("cut")
    cutting.append(["数量", "备件编码", "备件名称", "库位"])
    cutting.append([6, "C.G.LM.000011000", "Robot", "A-B03-011"])
    workbook.save(path)

    from backend.app.services import outbound_reconciliation
    from backend.app.services.material_mapping import MaterialMatch

    class FakeSettings:
        outbound_workbook_file = path
        outbound_shipping_sheet = "ship"
        outbound_cutting_sheet = "cut"
        outbound_cutting_text_file = tmp_path / "missing-cut.txt"
        outbound_shipping_text_file = tmp_path / "missing-ship.txt"

    monkeypatch.setattr(outbound_reconciliation, "get_settings", lambda: FakeSettings())
    monkeypatch.setattr(
        outbound_reconciliation,
        "find_material_matches",
        lambda code: [
            MaterialMatch(
                ruiyun_part_number="C.G.LM.000011000",
                sku="MTL24LUM1US02",
                matched_input=code,
                matched_field="sku",
            )
        ],
    )

    result = query_outbound("MTL24LUM1US02", selected_orders=["SO202604210135"])

    assert result["belongs_to_selected"] is True
    assert result["shipping_orders"][0]["part_code"] == "C.G.LM.000011000"
