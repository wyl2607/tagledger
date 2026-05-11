from pathlib import Path

from openpyxl import Workbook

from backend.app.services.material_mapping import (
    find_material_matches,
    load_material_matches,
    normalize_material_code,
)


def write_mapping(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["物料编码", "SKU", "BYD料号", "瑞云料号", "描述", "类型", "品名", "箱规尺寸"])
    sheet.append(
        [
            "MAT-001",
            "MTL24LUM1US02",
            "BYD-001",
            "C.G.LM.000011000",
            "后轮电机控制器",
            "电控",
            "控制器",
            "40x30x20",
        ]
    )
    sheet.append(
        [
            "MAT-002",
            "MTL24LUL1EU02",
            "BYD-002",
            "C.G.LD.000010000",
            "前轮灯组",
            "灯具",
            "灯组",
            "20x10x8",
        ]
    )
    workbook.save(path)


def test_load_material_matches_reads_ruiyun_and_sku_columns(tmp_path: Path) -> None:
    path = tmp_path / "mapping.xlsx"
    write_mapping(path)

    matches = load_material_matches(path)

    assert matches[0].ruiyun_part_number == "C.G.LM.000011000"
    assert matches[0].sku == "MTL24LUM1US02"
    assert matches[0].material_code == "MAT-001"
    assert matches[0].byd_part_number == "BYD-001"
    assert matches[0].description == "后轮电机控制器"
    assert matches[0].material_type == "电控"
    assert matches[0].product_name == "控制器"
    assert matches[0].carton_size == "40x30x20"


def test_normalize_material_code_ignores_punctuation() -> None:
    assert normalize_material_code(" C.G.LM.000011000 ") == "CGLM000011000"


def test_find_material_matches_returns_both_numbers(tmp_path: Path, monkeypatch) -> None:
    path = tmp_path / "mapping.xlsx"
    write_mapping(path)

    from backend.app.services import material_mapping

    material_mapping.clear_material_mapping_cache()

    class FakeSettings:
        material_mapping_file = path

    monkeypatch.setattr(material_mapping, "get_settings", lambda: FakeSettings())

    matches = find_material_matches("OCR text SKU MTL24LUM1US02")

    assert len(matches) == 1
    assert matches[0].ruiyun_part_number == "C.G.LM.000011000"
    assert matches[0].sku == "MTL24LUM1US02"
    assert matches[0].matched_field == "sku"


def test_find_material_matches_accepts_byd_and_material_code(tmp_path: Path, monkeypatch) -> None:
    path = tmp_path / "mapping.xlsx"
    write_mapping(path)

    from backend.app.services import material_mapping

    material_mapping.clear_material_mapping_cache()

    class FakeSettings:
        material_mapping_file = path

    monkeypatch.setattr(material_mapping, "get_settings", lambda: FakeSettings())

    by_byd = find_material_matches("扫码 BYD-001")
    by_code = find_material_matches("扫码 MAT-002")

    assert by_byd[0].matched_field == "byd_part_number"
    assert by_code[0].matched_field == "material_code"


def test_search_material_catalog_defaults_to_all_rows(tmp_path: Path, monkeypatch) -> None:
    path = tmp_path / "mapping.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["物料编码", "SKU", "BYD料号", "瑞云料号", "描述", "类型", "品名", "箱规尺寸"])
    for index in range(35):
        sheet.append(
            [
                f"MAT-{index:03d}",
                f"SKU-{index:03d}",
                f"BYD-{index:03d}",
                f"RY-{index:03d}",
                f"描述 {index}",
                "类型A" if index == 34 else "类型B",
                f"品名 {index}",
                f"{index}x10x8",
            ]
        )
    workbook.save(path)

    from backend.app.services import material_mapping

    material_mapping.clear_material_mapping_cache()

    class FakeSettings:
        material_mapping_file = path

    monkeypatch.setattr(material_mapping, "get_settings", lambda: FakeSettings())

    assert len(material_mapping.search_material_catalog()) == 35
    assert material_mapping.search_material_catalog("SKU-034")[0].sku == "SKU-034"
    assert material_mapping.search_material_catalog("类型A")[0].material_code == "MAT-034"


def test_load_material_matches_skips_unknown_headers(tmp_path: Path) -> None:
    path = tmp_path / "mapping.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["序号", "名称"])
    sheet.append(["1", "MTL24LUM1US02"])
    workbook.save(path)

    assert load_material_matches(path) == []


def test_material_catalog_refreshes_when_file_changes(tmp_path: Path, monkeypatch) -> None:
    path = tmp_path / "mapping.xlsx"
    write_mapping(path)

    from backend.app.services import material_mapping

    material_mapping.clear_material_mapping_cache()

    class FakeSettings:
        material_mapping_file = path

    monkeypatch.setattr(material_mapping, "get_settings", lambda: FakeSettings())
    assert find_material_matches("MTL24LUM1US02")[0].ruiyun_part_number == "C.G.LM.000011000"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["瑞云料号", "SKU"])
    sheet.append(["C.NEW.0001", "MTL24LUM1US02"])
    workbook.save(path)

    assert find_material_matches("MTL24LUM1US02")[0].ruiyun_part_number == "C.NEW.0001"
