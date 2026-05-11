import csv
import json
import re
import socket
import uuid
from collections import defaultdict
from dataclasses import dataclass
from datetime import UTC, datetime, timedelta
from io import StringIO
from pathlib import Path
from typing import NoReturn

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.exc import IntegrityError
from sqlmodel import Session, func, select

from backend.app.config import get_settings
from backend.app.models import (
    AuditLog,
    InventoryLocation,
    InventoryMovement,
    OutboundProgressSnapshot,
    OutboundScan,
    Record,
    User,
)
from backend.app.services.material_mapping import (
    find_material_matches,
    normalize_material_code,
)

ORDER_RE = re.compile(r"[S5]\s*[O0](?:\D*\d){10,12}", re.IGNORECASE)
PART_CODE_RE = re.compile(
    r"(?:[A-Z]\s*\.\s*){1,3}[A-Z0-9]{1,4}\s*\.\s*\d{4,9}[A-Z]?",
    re.IGNORECASE,
)
QTY_RE = re.compile(r"\d{1,5}")


@dataclass(frozen=True)
class OutboundItem:
    order_no: str
    part_code: str
    quantity: int | None
    source: str
    raw_line: str
    name: str = ""
    locations: tuple[str, ...] = ()


@dataclass(frozen=True)
class ReconciledItem:
    order_no: str
    part_code: str
    cutting_qty: int
    shipping_qty: int
    difference: int
    status: str
    cutting_lines: list[str]
    shipping_lines: list[str]


@dataclass(frozen=True)
class OutboundCompletionMark:
    order_no: str
    part_code: str
    quantity: int | None
    raw_line: str


class OutboundVerificationRequiredError(RuntimeError):
    pass


def normalize_order_no(value: str) -> str:
    cleaned = "".join(ch for ch in value.upper() if ch.isalnum())
    if cleaned.startswith("5"):
        cleaned = "S" + cleaned[1:]
    return cleaned


def normalize_order_set(values: list[str] | None) -> set[str]:
    return {normalize_order_no(value) for value in values or [] if normalize_order_no(value)}


def normalize_part_code(value: str) -> str:
    compact = re.sub(r"\s+", "", value.upper())
    compact = re.sub(r"\.+", ".", compact)
    return compact.strip(".,;:|[](){}")


def compact_part_code(value: str) -> str:
    return normalize_material_code(normalize_part_code(value))


def _quantity_before(line: str, code_start: int) -> int | None:
    prefix = ORDER_RE.sub(" ", line[:code_start])
    candidates = [int(match.group(0)) for match in QTY_RE.finditer(prefix)]
    return candidates[-1] if candidates else None


def _has_quantity_completion_mark(
    line: str, order_end: int, code_start: int
) -> tuple[bool, int | None]:
    between = line[order_end:code_start]
    quantity_matches = list(QTY_RE.finditer(between))
    if not quantity_matches:
        return False, None
    quantity_match = quantity_matches[-1]
    after_quantity = between[quantity_match.end() :]
    tokens = [
        token for token in re.split(r"[\s|_\-.:;,'\[\](){}]+", after_quantity.strip()) if token
    ]
    if not tokens:
        return False, int(quantity_match.group(0))
    # OCR commonly reads handwritten crosses/checks as K, V, ¥ or stray percent-like marks.
    return len(tokens[0]) == 1 and tokens[0].upper() in {
        "X",
        "×",
        "✕",
        "✖",
        "√",
        "✓",
        "V",
        "K",
        "¥",
        "%",
    }, int(quantity_match.group(0))


def parse_outbound_text(text: str, source: str) -> list[OutboundItem]:
    items: list[OutboundItem] = []
    current_order = ""
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        order_match = ORDER_RE.search(line)
        if order_match:
            current_order = normalize_order_no(order_match.group(0))
        if not current_order:
            continue
        for code_match in PART_CODE_RE.finditer(line):
            part_code = normalize_part_code(code_match.group(0))
            if part_code.count(".") < 2:
                continue
            items.append(
                OutboundItem(
                    order_no=current_order,
                    part_code=part_code,
                    quantity=_quantity_before(line, code_match.start()),
                    source=source,
                    raw_line=line,
                )
            )
    return items


def parse_outbound_completion_marks(text: str) -> list[OutboundCompletionMark]:
    marks: list[OutboundCompletionMark] = []
    current_order = ""
    seen: set[tuple[str, str, str]] = set()
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        order_match = ORDER_RE.search(line)
        if order_match:
            current_order = normalize_order_no(order_match.group(0))
            order_end = order_match.end()
        else:
            order_end = 0
        if not current_order:
            continue
        for code_match in PART_CODE_RE.finditer(line):
            part_code = normalize_part_code(code_match.group(0))
            if part_code.count(".") < 2:
                continue
            marked, quantity = _has_quantity_completion_mark(line, order_end, code_match.start())
            if not marked:
                continue
            key = (current_order, compact_part_code(part_code), line)
            if key in seen:
                continue
            seen.add(key)
            marks.append(
                OutboundCompletionMark(
                    order_no=current_order,
                    part_code=part_code,
                    quantity=quantity,
                    raw_line=line,
                )
            )
    return marks


def _read_text(path: Path) -> str:
    if not path.exists():
        return ""
    return path.read_text(encoding="utf-8", errors="ignore")


def _cell_text(value: object) -> str:
    return str(value).strip() if value is not None else ""


def _cell_int(value: object) -> int | None:
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = _cell_text(value)
    return int(text) if text.isdigit() else None


def _has_visible_mark(value: object) -> bool:
    text = _cell_text(value).upper()
    return text in {"X", "×", "√", "Y", "YES", "OK", "DONE", "已剪"}


def _is_invalid_location(value: str) -> bool:
    text = value.strip().upper()
    return text in {"#VALUE!", "#N/A", "#REF!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!"}


def _row_locations(row: tuple[object, ...], *, start_index: int = 3) -> tuple[str, ...]:
    locations = []
    for cell in row[start_index:]:
        text = _cell_text(cell)
        if text and not _has_visible_mark(text) and not _is_invalid_location(text):
            locations.append(text)
    return tuple(locations)


def _normalize_header_text(value: object) -> str:
    text = _cell_text(value)
    text = text.replace("（", "(").replace("）", ")")
    return re.sub(r"\s+", "", text).lower()


def _resolve_column_indexes(sheet) -> dict[str, int]:
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header):
        token = _normalize_header_text(cell)
        if not token:
            continue
        if "出库单号" in token and "order_no" not in mapping:
            mapping["order_no"] = idx
        if ("数量" in token or "qty" in token or "quantity" in token) and "quantity" not in mapping:
            mapping["quantity"] = idx
        if (
            "备件编码" in token or "编码" in token or "part" in token or "sku" in token
        ) and "part_code" not in mapping:
            mapping["part_code"] = idx
        if ("备件名称" in token or "名称" in token or "name" in token) and "name" not in mapping:
            mapping["name"] = idx
        if (
            "库位" in token or "location" in token or "loc" in token
        ) and "location_start" not in mapping:
            mapping["location_start"] = idx
    return mapping


def _load_shipping_sheet(path: Path, sheet_name: str) -> list[OutboundItem]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return []
        sheet = workbook[sheet_name]
        items: list[OutboundItem] = []
        col = _resolve_column_indexes(sheet)
        order_idx = col.get("order_no", 0)
        qty_idx = col.get("quantity", 1)
        part_idx = col.get("part_code", 2)
        name_idx = col.get("name", 3)
        current_order_no = ""
        for row in sheet.iter_rows(min_row=2, values_only=True):
            order_no = normalize_order_no(
                _cell_text(row[order_idx] if len(row) > order_idx else "")
            )
            if order_no:
                current_order_no = order_no
            else:
                order_no = current_order_no
            quantity = _cell_int(row[qty_idx] if len(row) > qty_idx else None)
            part_code = normalize_part_code(
                _cell_text(row[part_idx] if len(row) > part_idx else "")
            )
            name = _cell_text(row[name_idx] if len(row) > name_idx else "")
            if order_no and part_code:
                items.append(
                    OutboundItem(
                        order_no=order_no,
                        part_code=part_code,
                        quantity=quantity,
                        source="shipping",
                        raw_line=f"{order_no} {quantity or ''} {part_code} {name}".strip(),
                        name=name,
                    )
                )
        return items
    finally:
        workbook.close()


def _load_cutting_sheet(path: Path, sheet_name: str) -> list[OutboundItem]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return []
        sheet = workbook[sheet_name]
        items: list[OutboundItem] = []
        col = _resolve_column_indexes(sheet)
        order_idx = col.get("order_no")
        qty_idx = col.get("quantity", 0)
        part_idx = col.get("part_code", 1)
        name_idx = col.get("name", 2)
        location_start = col.get("location_start", 3)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            order_no = (
                normalize_order_no(_cell_text(row[order_idx] if len(row) > order_idx else ""))
                if order_idx is not None
                else ""
            )
            quantity = _cell_int(row[qty_idx] if len(row) > qty_idx else None)
            part_code = normalize_part_code(
                _cell_text(row[part_idx] if len(row) > part_idx else "")
            )
            name = _cell_text(row[name_idx] if len(row) > name_idx else "")
            checked = any(_has_visible_mark(cell) for cell in row[location_start:])
            locations = _row_locations(tuple(row), start_index=location_start)
            if part_code and ((quantity is not None and quantity > 0) or checked):
                items.append(
                    OutboundItem(
                        order_no=order_no or "PICKING_TOTAL",
                        part_code=part_code,
                        quantity=quantity,
                        source="cutting",
                        raw_line=f"{quantity or ''} {part_code} {name}".strip(),
                        name=name,
                        locations=locations,
                    )
                )
        return items
    finally:
        workbook.close()


def _real_order_numbers(items: list[OutboundItem]) -> set[str]:
    return {item.order_no for item in items if item.order_no and item.order_no != "PICKING_TOTAL"}


def _shipping_items_from_ordered_cutting(items: list[OutboundItem]) -> list[OutboundItem]:
    return [
        OutboundItem(
            order_no=item.order_no,
            part_code=item.part_code,
            quantity=item.quantity,
            source="shipping",
            raw_line=item.raw_line,
            name=item.name,
            locations=item.locations,
        )
        for item in items
        if item.order_no != "PICKING_TOTAL"
    ]


def _fail_outbound_config(message: str) -> NoReturn:
    raise RuntimeError(message)


def load_outbound_items() -> tuple[list[OutboundItem], list[OutboundItem]]:
    settings = get_settings()
    if not settings.outbound_workbook_file.exists():
        _fail_outbound_config(f"outbound workbook not found: {settings.outbound_workbook_file}")
    cutting = _load_cutting_sheet(settings.outbound_workbook_file, settings.outbound_cutting_sheet)
    shipping = _load_shipping_sheet(
        settings.outbound_workbook_file, settings.outbound_shipping_sheet
    )
    cutting_orders = _real_order_numbers(cutting)
    if cutting_orders:
        scoped_shipping = [
            item for item in shipping if normalize_order_no(item.order_no) in cutting_orders
        ]
        shipping = scoped_shipping or _shipping_items_from_ordered_cutting(cutting)
    if not cutting or not shipping:
        _fail_outbound_config(
            f"outbound workbook missing required rows or sheets: {settings.outbound_workbook_file}"
        )
    return cutting, shipping


def _aggregate(items: list[OutboundItem]) -> dict[tuple[str, str], dict[str, object]]:
    grouped: dict[tuple[str, str], dict[str, object]] = {}
    for item in items:
        key = (item.order_no, compact_part_code(item.part_code))
        if key not in grouped:
            grouped[key] = {
                "order_no": item.order_no,
                "part_code": item.part_code,
                "quantity": 0,
                "lines": [],
                "names": [],
                "unknown_quantity": False,
            }
        if item.quantity is None:
            grouped[key]["unknown_quantity"] = True
        else:
            grouped[key]["quantity"] = int(grouped[key]["quantity"]) + item.quantity
        grouped[key]["lines"].append(item.raw_line)
        if item.name:
            grouped[key]["names"].append(item.name)
        if item.locations:
            grouped[key].setdefault("locations", [])
            grouped[key]["locations"].extend(item.locations)
    return grouped


def _aggregate_by_part(items: list[OutboundItem]) -> dict[str, dict[str, object]]:
    grouped: dict[str, dict[str, object]] = {}
    for item in items:
        key = compact_part_code(item.part_code)
        if key not in grouped:
            grouped[key] = {
                "part_code": item.part_code,
                "quantity": 0,
                "lines": [],
                "names": [],
                "locations": [],
                "unknown_quantity": False,
            }
        if item.quantity is not None:
            grouped[key]["quantity"] = int(grouped[key]["quantity"]) + item.quantity
        else:
            grouped[key]["unknown_quantity"] = True
        grouped[key]["lines"].append(item.raw_line)
        if item.name:
            grouped[key]["names"].append(item.name)
        grouped[key]["locations"].extend(item.locations)
    return grouped


def reconcile_outbound_items(
    cutting_items: list[OutboundItem],
    shipping_items: list[OutboundItem],
) -> list[ReconciledItem]:
    cutting = _aggregate(cutting_items)
    shipping = _aggregate(shipping_items)
    rows: list[ReconciledItem] = []
    for key in sorted(set(cutting) | set(shipping)):
        cut = cutting.get(key)
        ship = shipping.get(key)
        cutting_qty = int(cut["quantity"]) if cut else 0
        shipping_qty = int(ship["quantity"]) if ship else 0
        if cut is None:
            status = "shipping_extra"
        elif ship is None:
            status = "cutting_missing_shipping"
        elif cut.get("unknown_quantity") or ship.get("unknown_quantity"):
            status = "quantity_unreadable"
        elif cutting_qty == shipping_qty:
            status = "matched"
        elif shipping_qty > cutting_qty:
            status = "over_shipped"
        else:
            status = "under_shipped"
        rows.append(
            ReconciledItem(
                order_no=(cut or ship)["order_no"],
                part_code=(cut or ship)["part_code"],
                cutting_qty=cutting_qty,
                shipping_qty=shipping_qty,
                difference=shipping_qty - cutting_qty,
                status=status,
                cutting_lines=list(cut["lines"]) if cut else [],
                shipping_lines=list(ship["lines"]) if ship else [],
            )
        )
    return rows


def _part_rows(
    cutting_items: list[OutboundItem], shipping_items: list[OutboundItem]
) -> list[dict[str, object]]:
    cutting_parts = _aggregate_by_part(cutting_items)
    shipping_parts = _aggregate_by_part(shipping_items)
    rows = []
    for key in sorted(set(cutting_parts) | set(shipping_parts)):
        cut = cutting_parts.get(key)
        ship = shipping_parts.get(key)
        cutting_qty = int(cut["quantity"]) if cut else 0
        shipping_qty = int(ship["quantity"]) if ship else 0
        if cut is None:
            status = "shipping_extra"
        elif ship is None:
            status = "cutting_missing_shipping"
        elif cut.get("unknown_quantity") or ship.get("unknown_quantity"):
            status = "quantity_unreadable"
        elif cutting_qty == shipping_qty:
            status = "matched"
        elif shipping_qty > cutting_qty:
            status = "over_shipped"
        else:
            status = "under_shipped"
        rows.append(
            {
                "part_code": (cut or ship)["part_code"],
                "name": next(iter((cut or ship).get("names", [])), ""),
                "cutting_qty": cutting_qty,
                "shipping_qty": shipping_qty,
                "difference": shipping_qty - cutting_qty,
                "status": status,
                "cutting_lines": list(cut["lines"]) if cut else [],
                "shipping_lines": list(ship["lines"]) if ship else [],
                "locations": sorted(set(cut.get("locations", []))) if cut else [],
            }
        )
    return rows


def _scope_items_for_orders(
    cutting_items: list[OutboundItem],
    shipping_items: list[OutboundItem],
    allowed_orders: list[str] | None,
) -> tuple[list[OutboundItem], list[OutboundItem]]:
    if allowed_orders is None:
        return cutting_items, shipping_items
    selected = normalize_order_set(allowed_orders)
    if not selected:
        return [], []
    scoped_shipping = [
        item for item in shipping_items if normalize_order_no(item.order_no) in selected
    ]
    scoped_part_keys = {compact_part_code(item.part_code) for item in scoped_shipping}
    scoped_cutting = [
        item for item in cutting_items if compact_part_code(item.part_code) in scoped_part_keys
    ]
    return scoped_cutting, scoped_shipping


def outbound_summary(allowed_orders: list[str] | None = None) -> dict[str, object]:
    cutting_items, shipping_items = load_outbound_items()
    cutting_items, shipping_items = _scope_items_for_orders(
        cutting_items,
        shipping_items,
        allowed_orders,
    )
    part_rows = _part_rows(cutting_items, shipping_items)
    part_status_counts: dict[str, int] = defaultdict(int)
    for row in part_rows:
        part_status_counts[str(row["status"])] += 1
    order_numbers = {
        "cutting": sorted({item.order_no for item in cutting_items}),
        "shipping": sorted({item.order_no for item in shipping_items}),
    }
    return {
        "data_source": "workbook" if get_settings().outbound_workbook_file.exists() else "text_ocr",
        "order_numbers": {
            "shipping": order_numbers["shipping"],
            "note": "拣货单没有出库单号，单号只来自发货单；数量按备件编码总量核对。",
        },
        "counts": {
            "cutting_items": len(cutting_items),
            "shipping_items": len(shipping_items),
            "cutting_part_rows": len(_aggregate_by_part(cutting_items)),
            "shipping_part_rows": len(_aggregate_by_part(shipping_items)),
            "part_status": dict(sorted(part_status_counts.items())),
        },
        "part_rows": part_rows,
    }


def outbound_order_choices(allowed_orders: list[str] | None = None) -> dict[str, object]:
    cutting_items, shipping_items = load_outbound_items()
    _, shipping_items = _scope_items_for_orders(cutting_items, shipping_items, allowed_orders)
    return {
        "data_source": "workbook" if get_settings().outbound_workbook_file.exists() else "text_ocr",
        "order_numbers": {
            "shipping": sorted(
                {normalize_order_no(item.order_no) for item in shipping_items if item.order_no}
            ),
            "note": "拣货单没有出库单号，单号只来自发货单；数量按备件编码总量核对。",
        },
    }


def query_outbound(
    code: str,
    selected_orders: list[str] | None = None,
    allowed_orders: list[str] | None = None,
) -> dict[str, object]:
    candidates = {compact_part_code(code)}
    material_matches = []
    for match in find_material_matches(code):
        material_matches.append(match.__dict__)
        candidates.add(compact_part_code(match.ruiyun_part_number))
        candidates.add(compact_part_code(match.sku))
    cutting_items, shipping_items = load_outbound_items()
    cutting_items, shipping_items = _scope_items_for_orders(
        cutting_items,
        shipping_items,
        allowed_orders,
    )
    rows = [
        row
        for row in _part_rows(cutting_items, shipping_items)
        if compact_part_code(str(row["part_code"])) in candidates
    ]
    shipping_orders = [
        item.__dict__ for item in shipping_items if compact_part_code(item.part_code) in candidates
    ]
    cutting_totals = [
        item.__dict__ for item in cutting_items if compact_part_code(item.part_code) in candidates
    ]
    selected = normalize_order_set(selected_orders)
    matching_selected_orders = [
        item for item in shipping_orders if normalize_order_no(str(item["order_no"])) in selected
    ]
    matching_other_orders = [
        item
        for item in shipping_orders
        if selected and normalize_order_no(str(item["order_no"])) not in selected
    ]
    return {
        "query": code,
        "candidate_codes": sorted(candidates),
        "material_matches": material_matches,
        "selected_order_numbers": sorted(selected),
        "belongs_to_selected": bool(selected and matching_selected_orders),
        "rows": rows,
        "shipping_orders": shipping_orders,
        "matching_selected_orders": matching_selected_orders,
        "matching_other_orders": matching_other_orders,
        "cutting_totals": cutting_totals,
    }


def candidate_part_codes(code: str) -> list[str]:
    candidates = {compact_part_code(code)}
    for match in find_material_matches(code):
        candidates.add(compact_part_code(match.ruiyun_part_number))
        candidates.add(compact_part_code(match.sku))
    return sorted(candidate for candidate in candidates if candidate)


def _order_required_rows(order_no: str) -> dict[str, dict[str, object]]:
    selected_order = normalize_order_no(order_no)
    cutting_items, shipping_items = load_outbound_items()
    location_map = _part_location_map(cutting_items)
    rows: dict[str, dict[str, object]] = {}
    for item in shipping_items:
        if normalize_order_no(item.order_no) != selected_order:
            continue
        key = compact_part_code(item.part_code)
        if key not in rows:
            rows[key] = {
                "order_no": selected_order,
                "part_code": item.part_code,
                "name": item.name,
                "locations": location_map.get(key, []),
                "required_qty": 0,
                "unknown_quantity": False,
                "shipping_lines": [],
            }
        if item.quantity is None:
            rows[key]["unknown_quantity"] = True
        else:
            rows[key]["required_qty"] = int(rows[key]["required_qty"]) + item.quantity
        rows[key]["shipping_lines"].append(item.raw_line)
    return rows


def _order_numbers() -> list[str]:
    _, shipping_items = load_outbound_items()
    return sorted({normalize_order_no(item.order_no) for item in shipping_items if item.order_no})


def _part_location_map(cutting_items: list[OutboundItem]) -> dict[str, list[str]]:
    locations: dict[str, list[str]] = defaultdict(list)
    for item in cutting_items:
        key = compact_part_code(item.part_code)
        for location in item.locations:
            if location not in locations[key]:
                locations[key].append(location)
    return dict(locations)


def _today_part_remaining(session: Session) -> dict[str, int]:
    orders = [
        outbound_order_status(order_no, session, include_today_remaining=False)
        for order_no in _order_numbers()
    ]
    totals: dict[str, int] = defaultdict(int)
    for order in orders:
        for row in order["rows"]:
            totals[str(row["part_key"])] += int(row.get("remaining_qty") or 0)
    return dict(totals)


def _scan_counts(session: Session, order_no: str) -> dict[str, int]:
    selected_order = normalize_order_no(order_no)
    statement = (
        select(OutboundScan.part_code, func.sum(OutboundScan.quantity))
        .where(OutboundScan.order_no == selected_order, OutboundScan.status == "active")
        .group_by(OutboundScan.part_code)
    )
    return {
        str(part_code): int(quantity or 0) for part_code, quantity in session.exec(statement).all()
    }


def _scan_total(session: Session, order_no: str) -> int:
    selected_order = normalize_order_no(order_no)
    statement = select(func.sum(OutboundScan.quantity)).where(
        OutboundScan.order_no == selected_order,
        OutboundScan.status == "active",
    )
    return int(session.exec(statement).one() or 0)


def _active_scan_summary(session: Session, order_no: str) -> tuple[int, int]:
    selected_order = normalize_order_no(order_no)
    statement = select(func.count(OutboundScan.id), func.sum(OutboundScan.quantity)).where(
        OutboundScan.order_no == selected_order,
        OutboundScan.status == "active",
    )
    count, quantity = session.exec(statement).one()
    return int(count or 0), int(quantity or 0)


def _is_same_batch_scope(scan_batch_id: str | None, batch_id: str | None) -> bool:
    current = (scan_batch_id or "").strip()
    target = (batch_id or "").strip()
    return current == target if target else current == ""


def _snapshot_to_payload(snapshot: OutboundProgressSnapshot) -> dict[str, object]:
    detail = json.loads(snapshot.detail_json) if snapshot.detail_json else {}
    detail.pop("source_code", None)
    return {
        "id": snapshot.id,
        "order_no": snapshot.order_no,
        "event": snapshot.event,
        "required_total": snapshot.required_total,
        "scanned_total": snapshot.scanned_total,
        "remaining_total": snapshot.remaining_total,
        "line_total": snapshot.line_total,
        "complete_line_total": snapshot.complete_line_total,
        "active_scan_count": snapshot.active_scan_count,
        "active_scan_quantity": snapshot.active_scan_quantity,
        "operator_id": snapshot.operator_id,
        "batch_id": snapshot.batch_id,
        "scan_id": snapshot.scan_id,
        "completed_at": snapshot.completed_at.isoformat() if snapshot.completed_at else None,
        "detail": detail,
        "created_at": snapshot.created_at.isoformat() if snapshot.created_at else None,
    }


def save_outbound_progress_snapshot(
    *,
    order_no: str,
    event: str,
    operator_id: str,
    session: Session,
    status: dict[str, object] | None = None,
    scan_id: int | None = None,
    batch_id: str | None = None,
    detail: dict[str, object] | None = None,
    completed_at: datetime | None = None,
) -> OutboundProgressSnapshot:
    selected_order = normalize_order_no(order_no)
    status = status or outbound_order_status(selected_order, session)
    active_count, active_quantity = _active_scan_summary(session, selected_order)
    snapshot = OutboundProgressSnapshot(
        order_no=selected_order,
        event=event[:80],
        required_total=int(status.get("required_total") or 0),
        scanned_total=int(status.get("scanned_total") or 0),
        remaining_total=int(status.get("remaining_total") or 0),
        line_total=int(status.get("line_total") or 0),
        complete_line_total=int(status.get("complete_line_total") or 0),
        active_scan_count=active_count,
        active_scan_quantity=active_quantity,
        operator_id=(operator_id.strip() or "self")[:80],
        batch_id=(batch_id or "")[:120] or None,
        scan_id=scan_id,
        completed_at=completed_at,
        detail_json=json.dumps(detail or {}, ensure_ascii=False, default=str),
    )
    session.add(snapshot)
    session.commit()
    session.refresh(snapshot)
    return snapshot


def _manual_source_code(part_key: str) -> str:
    return f"MANUAL:{part_key}"


def _scan_to_payload(scan: OutboundScan) -> dict[str, object]:
    return {
        "id": scan.id,
        "order_no": scan.order_no,
        "part_code": scan.part_code,
        "location_code": scan.location_code,
        "source_code": scan.source_code,
        "matched_code": scan.matched_code,
        "quantity": scan.quantity,
        "status": scan.status,
        "operator_id": scan.operator_id,
        "batch_id": scan.batch_id,
        "record_id": scan.record_id,
        "verification_record_id": scan.verification_record_id,
        "void_reason": scan.void_reason,
        "voided_by": scan.voided_by,
        "created_at": scan.created_at.isoformat() if scan.created_at else None,
        "voided_at": scan.voided_at.isoformat() if scan.voided_at else None,
    }


def _source_part_key_from_scan_code(code: str) -> str | None:
    match = PART_CODE_RE.search(code)
    if match is None:
        return None
    return compact_part_code(normalize_part_code(match.group(0)))


def _normalize_location_code(value: str) -> str:
    cleaned = value.strip().upper()
    if not cleaned:
        raise RuntimeError("location_code is required")
    return cleaned


def _inventory_status_value(value: str) -> str:
    status = value.strip().lower()
    aliases = {
        "replenish_needed": "pending_restock",
        "to_be_replaced": "pending_replacement",
    }
    status = aliases.get(status, status)
    if status not in {
        "active",
        "disabled",
        "pending_restock",
        "pending_replacement",
        "zero_stock",
        "retired",
    }:
        raise RuntimeError(
            "status must be active, disabled, pending_restock, pending_replacement, zero_stock, or retired"
        )
    return status


def _normalize_location_kind(value: str | None) -> str:
    kind = (value or "").strip().lower()
    if kind not in {"permanent", "long_term", "temporary"}:
        return "permanent"
    return kind


def _get_or_create_inventory_location(
    session: Session,
    *,
    part_key: str,
    location_code: str,
) -> InventoryLocation:
    row = session.exec(
        select(InventoryLocation).where(
            InventoryLocation.part_key == part_key,
            InventoryLocation.location_code == location_code,
        )
    ).first()
    if row is not None:
        return row
    row = InventoryLocation(
        part_key=part_key,
        location_code=location_code,
        quantity=0,
        status="active",
        zero_stock=True,
    )
    session.add(row)
    session.commit()
    session.refresh(row)
    return row


def _find_inventory_location(
    session: Session,
    *,
    part_key: str,
    location_code: str,
) -> InventoryLocation | None:
    return session.exec(
        select(InventoryLocation).where(
            InventoryLocation.part_key == part_key,
            InventoryLocation.location_code == location_code,
        )
    ).first()


def _bootstrap_inventory_if_missing(
    session: Session,
    *,
    part_key: str,
    location_code: str,
    operator_id: str,
    reason: str,
    seed_quantity: int,
) -> InventoryLocation | None:
    normalized_part = compact_part_code(part_key)
    normalized_location = _normalize_location_code(location_code)
    existing = _find_inventory_location(
        session, part_key=normalized_part, location_code=normalized_location
    )
    if existing is not None:
        return None
    quantity = max(0, int(seed_quantity))
    row = InventoryLocation(
        part_key=normalized_part,
        location_code=normalized_location,
        quantity=quantity,
        status="active",
        zero_stock=quantity == 0,
    )
    session.add(row)
    session.commit()
    session.refresh(row)
    _record_inventory_movement(
        session,
        movement_type="bootstrap",
        part_key=normalized_part,
        location_code=normalized_location,
        order_no=None,
        scan_id=None,
        quantity_delta=quantity,
        before_qty=0,
        after_qty=quantity,
        operator_id=operator_id,
        reason=reason,
    )
    return row


def _list_order_locations(required_row: dict[str, object]) -> list[str]:
    return sorted(
        {
            _normalize_location_code(str(location))
            for location in required_row.get("locations", [])
            if str(location).strip()
        }
    )


def _select_location_for_outbound(
    *,
    location_code: str | None,
    required_row: dict[str, object],
) -> tuple[str | None, list[str], bool]:
    available_locations = _list_order_locations(required_row)
    requested = _normalize_location_code(location_code) if location_code else None
    if not available_locations:
        return requested, [], False
    if requested:
        return requested, available_locations, requested in available_locations
    if len(available_locations) == 1:
        return available_locations[0], available_locations, True
    return None, available_locations, False


def _inventory_location_payload(location: InventoryLocation) -> dict[str, object]:
    return {
        "id": location.id,
        "part_key": location.part_key,
        "part_name": location.part_name,
        "location_code": location.location_code,
        "quantity": int(location.quantity),
        "status": location.status,
        "zero_stock": bool(location.zero_stock),
        "location_kind": _normalize_location_kind(location.location_kind),
        "replacement_location_code": location.replacement_location_code,
        "updated_at": location.updated_at.isoformat() if location.updated_at else None,
    }


def _movement_payload(movement: InventoryMovement) -> dict[str, object]:
    return {
        "id": movement.id,
        "movement_type": movement.movement_type,
        "part_key": movement.part_key,
        "location_code": movement.location_code,
        "order_no": movement.order_no,
        "scan_id": movement.scan_id,
        "quantity_delta": movement.quantity_delta,
        "before_qty": movement.before_qty,
        "after_qty": movement.after_qty,
        "operator_id": movement.operator_id,
        "reason": movement.reason,
        "created_at": movement.created_at.isoformat() if movement.created_at else None,
    }


def _record_inventory_movement(
    session: Session,
    *,
    movement_type: str,
    part_key: str,
    location_code: str,
    order_no: str | None,
    scan_id: int | None,
    quantity_delta: int,
    before_qty: int,
    after_qty: int,
    operator_id: str,
    reason: str | None,
) -> InventoryMovement:
    row = InventoryMovement(
        movement_type=movement_type,
        part_key=part_key,
        location_code=location_code,
        order_no=order_no,
        scan_id=scan_id,
        quantity_delta=quantity_delta,
        before_qty=before_qty,
        after_qty=after_qty,
        operator_id=(operator_id.strip() or "self")[:80],
        reason=reason,
    )
    session.add(row)
    session.commit()
    session.refresh(row)
    return row


def _apply_inventory_delta(
    session: Session,
    *,
    movement_type: str,
    part_key: str,
    location_code: str,
    quantity_delta: int,
    operator_id: str,
    reason: str,
    order_no: str | None = None,
    scan_id: int | None = None,
    allow_new_location: bool = False,
) -> tuple[InventoryLocation, InventoryMovement]:
    normalized_part = compact_part_code(part_key)
    normalized_location = _normalize_location_code(location_code)
    location = _get_or_create_inventory_location(
        session, part_key=normalized_part, location_code=normalized_location
    )
    if (
        not allow_new_location
        and location.id is not None
        and location.quantity == 0
        and quantity_delta < 0
    ):
        raise RuntimeError(
            f"insufficient inventory at location {normalized_location} for part {normalized_part}"
        )
    if location.status == "disabled":
        raise RuntimeError(f"location {normalized_location} for part {normalized_part} is disabled")
    before_qty = int(location.quantity)
    after_qty = before_qty + quantity_delta
    if after_qty < 0:
        raise RuntimeError(
            f"insufficient inventory at location {normalized_location} for part {normalized_part}"
        )
    now = datetime.now(UTC)
    location.quantity = after_qty
    location.zero_stock = after_qty == 0
    location_kind = _normalize_location_kind(location.location_kind)
    if location.zero_stock and location_kind == "temporary":
        location.status = "retired"
    elif location.zero_stock:
        location.status = "zero_stock"
    elif location.status == "zero_stock":
        location.status = "active"
    elif location.status == "retired":
        location.status = "active"
    location.updated_at = now
    session.add(location)
    session.commit()
    session.refresh(location)
    movement = _record_inventory_movement(
        session,
        movement_type=movement_type,
        part_key=normalized_part,
        location_code=normalized_location,
        order_no=order_no,
        scan_id=scan_id,
        quantity_delta=quantity_delta,
        before_qty=before_qty,
        after_qty=after_qty,
        operator_id=operator_id,
        reason=reason,
    )
    return location, movement


def inbound_inventory(
    *,
    part_key: str,
    location_code: str,
    quantity: int,
    operator_id: str,
    reason: str,
    session: Session,
) -> dict[str, object]:
    qty = int(quantity)
    if qty <= 0:
        raise RuntimeError("quantity must be > 0")
    location, movement = _apply_inventory_delta(
        session,
        movement_type="inbound",
        part_key=part_key,
        location_code=location_code,
        quantity_delta=qty,
        operator_id=operator_id,
        reason=reason[:200] or "inbound",
        allow_new_location=True,
    )
    return {
        "updated": True,
        "location": _inventory_location_payload(location),
        "movement": _movement_payload(movement),
    }


def outbound_inventory(
    *,
    part_key: str,
    location_code: str,
    quantity: int,
    operator_id: str,
    reason: str,
    session: Session,
    order_no: str | None = None,
) -> dict[str, object]:
    qty = int(quantity)
    if qty <= 0:
        raise RuntimeError("quantity must be > 0")
    location, movement = _apply_inventory_delta(
        session,
        movement_type="outbound",
        part_key=part_key,
        location_code=location_code,
        quantity_delta=-qty,
        operator_id=operator_id,
        reason=reason[:200] or "outbound",
        order_no=(order_no or "").strip() or None,
    )
    return {
        "updated": True,
        "location": _inventory_location_payload(location),
        "movement": _movement_payload(movement),
    }


def transfer_inventory(
    *,
    part_key: str,
    from_location_code: str,
    to_location_code: str,
    quantity: int,
    operator_id: str,
    reason: str,
    session: Session,
) -> dict[str, object]:
    qty = int(quantity)
    if qty <= 0:
        raise RuntimeError("quantity must be > 0")
    normalized_part = compact_part_code(part_key)
    source_code = _normalize_location_code(from_location_code)
    target_code = _normalize_location_code(to_location_code)
    if source_code == target_code:
        raise RuntimeError("from_location_code and to_location_code must be different")

    source = _find_inventory_location(
        session,
        part_key=normalized_part,
        location_code=source_code,
    )
    if source is None or int(source.quantity) < qty:
        raise RuntimeError(
            f"insufficient inventory at location {source_code} for part {normalized_part}"
        )
    if source.status == "disabled":
        raise RuntimeError(f"location {source_code} for part {normalized_part} is disabled")

    target = _get_or_create_inventory_location(
        session,
        part_key=normalized_part,
        location_code=target_code,
    )
    if target.status == "disabled":
        raise RuntimeError(f"location {target_code} for part {normalized_part} is disabled")

    now = datetime.now(UTC)
    source_before = int(source.quantity)
    target_before = int(target.quantity)
    source.quantity = source_before - qty
    target.quantity = target_before + qty
    source.zero_stock = source.quantity == 0
    target.zero_stock = target.quantity == 0
    if source.zero_stock:
        source.status = "zero_stock"
    elif source.status == "zero_stock":
        source.status = "active"
    if target.status == "zero_stock":
        target.status = "active"
    source.updated_at = now
    target.updated_at = now
    session.add(source)
    session.add(target)
    session.commit()
    session.refresh(source)
    session.refresh(target)

    note = reason[:200] or "transfer"
    movement_out = _record_inventory_movement(
        session,
        movement_type="transfer_out",
        part_key=normalized_part,
        location_code=source_code,
        order_no=None,
        scan_id=None,
        quantity_delta=-qty,
        before_qty=source_before,
        after_qty=source_before - qty,
        operator_id=operator_id,
        reason=f"{note}; to={target_code}",
    )
    movement_in = _record_inventory_movement(
        session,
        movement_type="transfer_in",
        part_key=normalized_part,
        location_code=target_code,
        order_no=None,
        scan_id=None,
        quantity_delta=qty,
        before_qty=target_before,
        after_qty=target_before + qty,
        operator_id=operator_id,
        reason=f"{note}; from={source_code}",
    )
    return {
        "updated": True,
        "from_location": _inventory_location_payload(source),
        "to_location": _inventory_location_payload(target),
        "movements": [_movement_payload(movement_out), _movement_payload(movement_in)],
    }


def set_inventory_location_status(
    *,
    part_key: str,
    location_code: str,
    status: str,
    operator_id: str,
    reason: str,
    replacement_location_code: str | None,
    session: Session,
) -> dict[str, object]:
    normalized_part = compact_part_code(part_key)
    normalized_location = _normalize_location_code(location_code)
    target_status = _inventory_status_value(status)
    row = _get_or_create_inventory_location(
        session,
        part_key=normalized_part,
        location_code=normalized_location,
    )
    row.status = target_status
    row.replacement_location_code = (
        _normalize_location_code(replacement_location_code)
        if replacement_location_code and replacement_location_code.strip()
        else None
    )
    row.updated_at = datetime.now(UTC)
    row.zero_stock = int(row.quantity) <= 0
    session.add(row)
    session.commit()
    session.refresh(row)
    movement = _record_inventory_movement(
        session,
        movement_type="location_status",
        part_key=normalized_part,
        location_code=normalized_location,
        order_no=None,
        scan_id=None,
        quantity_delta=0,
        before_qty=int(row.quantity),
        after_qty=int(row.quantity),
        operator_id=operator_id,
        reason=(reason[:200] or "location_status"),
    )
    return {
        "updated": True,
        "location": _inventory_location_payload(row),
        "movement": _movement_payload(movement),
    }


def get_inventory_locations(
    session: Session,
    part_key: str | None = None,
    location_code: str | None = None,
    status: str | None = None,
    exclude_temporary: bool = False,
) -> dict[str, object]:
    statement = select(InventoryLocation).order_by(
        InventoryLocation.part_key.asc(),
        InventoryLocation.location_code.asc(),
    )
    normalized_part = compact_part_code(part_key) if part_key else ""
    normalized_location = _normalize_location_code(location_code) if location_code else ""
    normalized_status = _inventory_status_value(status) if status else ""
    if normalized_part:
        statement = statement.where(InventoryLocation.part_key == normalized_part)
    if normalized_location:
        statement = statement.where(InventoryLocation.location_code == normalized_location)
    if normalized_status:
        statement = statement.where(InventoryLocation.status == normalized_status)
    if exclude_temporary:
        statement = statement.where(
            func.lower(func.coalesce(InventoryLocation.location_kind, "permanent")) != "temporary"
        )
    rows = session.exec(statement).all()
    return {
        "part_key": normalized_part or None,
        "location_code": normalized_location or None,
        "status": normalized_status or None,
        "exclude_temporary": exclude_temporary,
        "locations": [_inventory_location_payload(row) for row in rows],
    }


def list_alternative_locations(
    session: Session,
    part_key: str,
    exclude_location: str | None = None,
) -> list[dict[str, object]]:
    normalized_part = compact_part_code(part_key)
    normalized_exclude = _normalize_location_code(exclude_location) if exclude_location else ""
    statement = (
        select(InventoryLocation)
        .where(InventoryLocation.part_key == normalized_part)
        .order_by(InventoryLocation.quantity.desc(), InventoryLocation.location_code.asc())
    )
    rows = session.exec(statement).all()
    alternatives: list[dict[str, object]] = []
    for row in rows:
        location_code = _normalize_location_code(row.location_code)
        quantity = int(row.quantity or 0)
        if quantity <= 0:
            continue
        if normalized_exclude and location_code == normalized_exclude:
            continue
        alternatives.append(
            {
                "location_code": location_code,
                "quantity": quantity,
                "status": row.status or "active",
                "location_kind": _normalize_location_kind(row.location_kind),
            }
        )
    return alternatives


def reactivate_inventory_location(
    *,
    inventory_location_id: int,
    reason: str,
    actor: User,
    session: Session,
) -> dict[str, object]:
    row = session.get(InventoryLocation, inventory_location_id)
    if row is None:
        raise RuntimeError(f"inventory location not found: {inventory_location_id}")
    if row.status != "retired":
        raise RuntimeError("only retired location can be reactivated")

    row.status = "active"
    row.updated_at = datetime.now(UTC)
    row.zero_stock = int(row.quantity) <= 0
    session.add(row)
    session.flush()

    movement = InventoryMovement(
        movement_type="location_reactivate",
        part_key=row.part_key,
        location_code=row.location_code,
        order_no=None,
        scan_id=None,
        quantity_delta=0,
        before_qty=int(row.quantity),
        after_qty=int(row.quantity),
        operator_id=actor.username,
        reason=(reason.strip() or "manual_reactivate")[:200],
    )
    session.add(movement)
    session.flush()

    audit = AuditLog(
        event_type="inventory_location_reactivate",
        actor_user_id=actor.id,
        actor_username=actor.username,
        target_type="inventory_location",
        target_id=str(row.id),
        action="inventory.location.reactivate",
        reason=movement.reason,
        success=True,
        detail_json=json.dumps(
            {
                "part_key": row.part_key,
                "location_code": row.location_code,
                "location_kind": _normalize_location_kind(row.location_kind),
                "from_status": "retired",
                "to_status": "active",
                "quantity": int(row.quantity),
            },
            ensure_ascii=False,
        ),
    )
    session.add(audit)
    session.commit()
    session.refresh(row)
    session.refresh(movement)
    session.refresh(audit)

    return {
        "updated": True,
        "location": _inventory_location_payload(row),
        "movement": _movement_payload(movement),
        "audit_log_id": audit.id,
    }


def get_inventory_movements(
    session: Session,
    *,
    part_key: str | None = None,
    location_code: str | None = None,
    limit: int = 100,
) -> dict[str, object]:
    statement = select(InventoryMovement).order_by(
        InventoryMovement.created_at.desc(),
        InventoryMovement.id.desc(),
    )
    normalized_part = compact_part_code(part_key) if part_key else ""
    normalized_location = _normalize_location_code(location_code) if location_code else ""
    if normalized_part:
        statement = statement.where(InventoryMovement.part_key == normalized_part)
    if normalized_location:
        statement = statement.where(InventoryMovement.location_code == normalized_location)
    rows = session.exec(statement.limit(max(1, min(limit, 500)))).all()
    return {
        "part_key": normalized_part or None,
        "location_code": normalized_location or None,
        "movements": [_movement_payload(row) for row in rows],
    }


def _safe_csv_cell(value: object) -> str:
    text = str(value or "")
    if text and text[0] in {"=", "+", "-", "@", "\t", "\r", "\n"}:
        return f"'{text}"
    return text


def _public_path(path: Path) -> str:
    return path.name


def outbound_order_status(
    order_no: str,
    session: Session,
    *,
    include_today_remaining: bool = True,
) -> dict[str, object]:
    selected_order = normalize_order_no(order_no)
    required_rows = _order_required_rows(selected_order)
    scan_counts = _scan_counts(session, selected_order)
    today_remaining = (
        _today_part_remaining(session) if include_today_remaining and required_rows else {}
    )
    rows = []
    for key, row in sorted(required_rows.items(), key=lambda item: str(item[1]["part_code"])):
        required_qty = int(row["required_qty"])
        scanned_qty = scan_counts.get(key, 0)
        remaining_qty = None if row["unknown_quantity"] else max(required_qty - scanned_qty, 0)
        over_scanned_qty = 0 if row["unknown_quantity"] else max(scanned_qty - required_qty, 0)
        rows.append(
            {
                **row,
                "part_key": key,
                "scanned_qty": scanned_qty,
                "remaining_qty": remaining_qty,
                "over_scanned_qty": over_scanned_qty,
                "today_remaining_qty": today_remaining.get(key, remaining_qty or 0),
                "is_complete": not row["unknown_quantity"] and scanned_qty >= required_qty,
            }
        )
    required_total = sum(int(row["required_qty"]) for row in required_rows.values())
    scanned_total = sum(scan_counts.get(key, 0) for key in required_rows)
    over_scanned_total = sum(int(row["over_scanned_qty"]) for row in rows)
    extra_scanned_total = (
        max(_scan_total(session, selected_order) - scanned_total, 0) + over_scanned_total
    )
    complete_line_total = sum(1 for row in rows if row["is_complete"])
    return {
        "order_no": selected_order,
        "required_total": required_total,
        "scanned_total": scanned_total,
        "extra_scanned_total": extra_scanned_total,
        "remaining_total": max(required_total - scanned_total, 0),
        "line_total": len(rows),
        "complete_line_total": complete_line_total,
        "is_complete": bool(rows) and complete_line_total == len(rows),
        "rows": rows,
    }


def outbound_orders_status(session: Session) -> dict[str, object]:
    orders = [
        outbound_order_status(order_no, session, include_today_remaining=False)
        for order_no in _order_numbers()
    ]
    today_remaining: dict[str, int] = defaultdict(int)
    for order in orders:
        for row in order["rows"]:
            today_remaining[str(row["part_key"])] += int(row.get("remaining_qty") or 0)
    for order in orders:
        for row in order["rows"]:
            row["today_remaining_qty"] = today_remaining.get(str(row["part_key"]), 0)
    complete_order_count = sum(1 for order in orders if order["is_complete"])
    return {
        "orders": orders,
        "totals": {
            "order_count": len(orders),
            "complete_order_count": complete_order_count,
            "open_order_count": len(orders) - complete_order_count,
            "required_total": sum(int(order["required_total"]) for order in orders),
            "scanned_total": sum(int(order["scanned_total"]) for order in orders),
            "remaining_total": sum(int(order["remaining_total"]) for order in orders),
            "extra_scanned_total": sum(int(order["extra_scanned_total"]) for order in orders),
        },
    }


def outbound_orders_overview(session: Session) -> dict[str, object]:
    status = outbound_orders_status(session)
    return {
        "orders": [
            {key: value for key, value in order.items() if key != "rows"}
            for order in status["orders"]
        ],
        "totals": status["totals"],
    }


def register_outbound_scan(
    *,
    order_no: str,
    code: str,
    operator_id: str,
    session: Session,
    record_id: int | None = None,
    verification_record_id: int | None = None,
    quantity: int = 1,
    location_code: str | None = None,
    dry_run: bool = False,
) -> dict[str, object]:
    selected_order = normalize_order_no(order_no)
    required_rows = _order_required_rows(selected_order)
    query_payload = query_outbound(code, selected_orders=[selected_order])
    quantity = max(1, int(quantity))
    if not required_rows:
        return {
            **query_payload,
            "scan_saved": False,
            "order_not_found": True,
            "order_status": outbound_order_status(selected_order, session),
        }
    candidates = candidate_part_codes(code)
    matched_key = next((candidate for candidate in candidates if candidate in required_rows), None)
    if matched_key is None:
        return {
            **query_payload,
            "scan_saved": False,
            "order_status": outbound_order_status(selected_order, session),
        }

    matched_row = required_rows[matched_key]
    source_part_key = _source_part_key_from_scan_code(code)
    requires_verification = bool(source_part_key and source_part_key != matched_key)
    if requires_verification and verification_record_id is None and not dry_run:
        raise OutboundVerificationRequiredError("verification_record_id is required")
    selected_location, available_locations, location_matches = _select_location_for_outbound(
        location_code=location_code,
        required_row=matched_row,
    )
    if available_locations and not selected_location:
        return {
            **query_payload,
            "scan_saved": False,
            "location_required": True,
            "available_locations": available_locations,
            "matched_part": matched_row,
            "order_status": outbound_order_status(selected_order, session),
        }
    if available_locations and not location_matches:
        return {
            **query_payload,
            "scan_saved": False,
            "location_invalid": True,
            "available_locations": available_locations,
            "requested_location_code": location_code,
            "matched_part": matched_row,
            "order_status": outbound_order_status(selected_order, session),
        }
    if record_id is not None:
        existing_statement = select(OutboundScan).where(
            OutboundScan.order_no == selected_order,
            OutboundScan.part_code == matched_key,
            OutboundScan.location_code == selected_location,
            OutboundScan.record_id == record_id,
            OutboundScan.status == "active",
        )
        if session.exec(existing_statement).first() is not None:
            return {
                **query_payload,
                "scan_saved": False,
                "already_recorded": True,
                "matched_part": matched_row,
                "order_status": outbound_order_status(selected_order, session),
            }
    before_counts = _scan_counts(session, selected_order)
    before_scanned = before_counts.get(matched_key, 0)
    required_qty = int(matched_row["required_qty"])
    unknown_quantity = bool(matched_row["unknown_quantity"])
    if unknown_quantity:
        return {
            **query_payload,
            "scan_saved": False,
            "quantity_unreadable": True,
            "matched_part": matched_row,
            "order_status": outbound_order_status(selected_order, session),
        }
    if not unknown_quantity and before_scanned >= required_qty:
        return {
            **query_payload,
            "scan_saved": False,
            "already_complete": True,
            "matched_part": matched_row,
            "order_status": outbound_order_status(selected_order, session),
        }
    remaining_qty = required_qty - before_scanned
    if quantity > remaining_qty:
        return {
            **query_payload,
            "scan_saved": False,
            "quantity_too_large": True,
            "requested_quantity": quantity,
            "remaining_qty": remaining_qty,
            "matched_part": matched_row,
            "order_status": outbound_order_status(selected_order, session),
        }
    if dry_run:
        return {
            **query_payload,
            "scan_saved": False,
            "preview": True,
            "location_code": selected_location,
            "requested_quantity": quantity,
            "remaining_qty": remaining_qty,
            "matched_part": matched_row,
            "requires_verification": requires_verification,
            "verification_reason": "part_mismatch" if requires_verification else None,
            "source_part_key": source_part_key,
            "order_status": outbound_order_status(selected_order, session),
        }
    inventory_location = None
    inventory_movement = None
    if selected_location:
        _bootstrap_inventory_if_missing(
            session,
            part_key=matched_key,
            location_code=selected_location,
            operator_id=operator_id,
            reason="legacy_bootstrap_from_order_requirement",
            seed_quantity=max(required_qty - before_scanned, quantity),
        )
        try:
            inventory_location, inventory_movement = _apply_inventory_delta(
                session,
                movement_type="outbound",
                part_key=matched_key,
                location_code=selected_location,
                quantity_delta=-quantity,
                operator_id=operator_id,
                reason="outbound_scan",
                order_no=selected_order,
            )
        except RuntimeError as exc:
            message = str(exc)
            if "disabled" in message:
                return {
                    **query_payload,
                    "scan_saved": False,
                    "location_disabled": True,
                    "location_code": selected_location,
                    "matched_part": matched_row,
                    "order_status": outbound_order_status(selected_order, session),
                }
            raise
    scan = OutboundScan(
        order_no=selected_order,
        part_code=matched_key,
        location_code=selected_location,
        source_code=code.strip(),
        matched_code=str(matched_row["part_code"]),
        quantity=quantity,
        status="active",
        operator_id=(operator_id.strip() or "self")[:80],
        record_id=record_id,
        verification_record_id=verification_record_id,
    )
    session.add(scan)
    try:
        session.commit()
    except IntegrityError:
        session.rollback()
        return {
            **query_payload,
            "scan_saved": False,
            "already_recorded": True,
            "matched_part": matched_row,
            "order_status": outbound_order_status(selected_order, session),
        }
    session.refresh(scan)
    if inventory_movement is not None:
        inventory_movement.scan_id = scan.id
        session.add(inventory_movement)
        session.commit()
        session.refresh(inventory_movement)
    status = outbound_order_status(selected_order, session)
    snapshot = save_outbound_progress_snapshot(
        order_no=selected_order,
        event="scan_confirmed",
        operator_id=operator_id,
        session=session,
        status=status,
        scan_id=scan.id,
        detail={"part_key": matched_key, "quantity": quantity, "source_code": code.strip()},
    )
    return {
        **query_payload,
        "scan_saved": True,
        "already_complete": False,
        "location_code": selected_location,
        "quantity": quantity,
        "requires_verification": requires_verification,
        "verification_reason": "part_mismatch" if requires_verification else None,
        "source_part_key": source_part_key,
        "matched_part": matched_row,
        "scan_id": scan.id,
        "scan": _scan_to_payload(scan),
        "inventory_location": _inventory_location_payload(inventory_location)
        if inventory_location is not None
        else None,
        "inventory_movement": _movement_payload(inventory_movement)
        if inventory_movement is not None
        else None,
        "snapshot": _snapshot_to_payload(snapshot),
        "order_status": status,
    }


def preview_outbound_scan(
    *,
    order_no: str,
    code: str,
    operator_id: str,
    session: Session,
    record_id: int | None = None,
    verification_record_id: int | None = None,
    quantity: int = 1,
    location_code: str | None = None,
) -> dict[str, object]:
    payload = register_outbound_scan(
        order_no=order_no,
        code=code,
        operator_id=operator_id,
        record_id=record_id,
        verification_record_id=verification_record_id,
        quantity=quantity,
        location_code=location_code,
        session=session,
        dry_run=True,
    )
    matched_part = payload.get("matched_part")
    matched_code = ""
    if isinstance(matched_part, dict):
        matched_code = str(matched_part.get("part_code") or "")
    selected_location = str(payload.get("location_code") or "").strip() or None
    if matched_code:
        payload["alternative_locations"] = list_alternative_locations(
            session=session,
            part_key=matched_code,
            exclude_location=selected_location,
        )
    else:
        payload["alternative_locations"] = []
    return payload


def outbound_order_scans(order_no: str, session: Session) -> dict[str, object]:
    selected_order = normalize_order_no(order_no)
    statement = (
        select(OutboundScan)
        .where(OutboundScan.order_no == selected_order)
        .order_by(OutboundScan.created_at.desc(), OutboundScan.id.desc())
    )
    return {
        "order_no": selected_order,
        "scans": [_scan_to_payload(scan) for scan in session.exec(statement).all()],
        "order_status": outbound_order_status(selected_order, session),
    }


def outbound_progress_snapshots(
    order_no: str | None,
    session: Session,
    *,
    limit: int = 100,
) -> dict[str, object]:
    statement = select(OutboundProgressSnapshot).order_by(
        OutboundProgressSnapshot.created_at.desc(),
        OutboundProgressSnapshot.id.desc(),
    )
    selected_order = normalize_order_no(order_no) if order_no else ""
    if selected_order:
        statement = statement.where(OutboundProgressSnapshot.order_no == selected_order)
    snapshots = session.exec(statement.limit(max(1, min(limit, 500)))).all()
    return {
        "order_no": selected_order or None,
        "snapshots": [_snapshot_to_payload(snapshot) for snapshot in snapshots],
    }


def outbound_remaining_csv(order_no: str, session: Session) -> str:
    status = outbound_order_status(order_no, session)
    buffer = StringIO()
    fields = [
        "order_no",
        "part_code",
        "name",
        "locations",
        "required_qty",
        "scanned_qty",
        "remaining_qty",
        "today_remaining_qty",
    ]
    writer = csv.DictWriter(buffer, fieldnames=fields)
    writer.writeheader()
    for row in status["rows"]:
        if int(row.get("remaining_qty") or 0) <= 0:
            continue
        writer.writerow(
            {
                "order_no": status["order_no"],
                "part_code": _safe_csv_cell(row["part_code"]),
                "name": _safe_csv_cell(row.get("name") or ""),
                "locations": _safe_csv_cell(
                    ";".join(str(location) for location in row.get("locations", []))
                ),
                "required_qty": row["required_qty"],
                "scanned_qty": row["scanned_qty"],
                "remaining_qty": row["remaining_qty"],
                "today_remaining_qty": row["today_remaining_qty"],
            }
        )
    return buffer.getvalue()


def outbound_batch_detail(order_no: str, batch_id: str, session: Session) -> dict[str, object]:
    selected_order = normalize_order_no(order_no)
    batch_id = batch_id.strip()
    if not batch_id:
        raise RuntimeError("batch_id is required")
    scans = session.exec(
        select(OutboundScan)
        .where(OutboundScan.order_no == selected_order, OutboundScan.batch_id == batch_id)
        .order_by(OutboundScan.created_at.desc(), OutboundScan.id.desc())
    ).all()
    if not scans:
        raise RuntimeError(f"outbound batch not found: {selected_order} {batch_id}")
    affected: dict[str, dict[str, object]] = {}
    for scan in scans:
        row = affected.setdefault(
            scan.part_code,
            {
                "part_key": scan.part_code,
                "matched_code": scan.matched_code,
                "active_quantity": 0,
                "voided_quantity": 0,
                "active_scan_count": 0,
                "voided_scan_count": 0,
            },
        )
        if scan.status == "active":
            row["active_quantity"] = int(row["active_quantity"]) + scan.quantity
            row["active_scan_count"] = int(row["active_scan_count"]) + 1
        else:
            row["voided_quantity"] = int(row["voided_quantity"]) + scan.quantity
            row["voided_scan_count"] = int(row["voided_scan_count"]) + 1
    active_scans = [scan for scan in scans if scan.status == "active"]
    active_quantity = sum(scan.quantity for scan in active_scans)
    status = outbound_order_status(selected_order, session)
    preview_status = {
        "scanned_total": max(int(status["scanned_total"]) - active_quantity, 0),
        "remaining_total": int(status["remaining_total"]) + active_quantity,
        "affected_quantity": active_quantity,
    }
    return {
        "order_no": selected_order,
        "batch_id": batch_id,
        "scan_count": len(scans),
        "active_scan_count": len(active_scans),
        "voided_scan_count": len(scans) - len(active_scans),
        "active_quantity": active_quantity,
        "voided_quantity": sum(scan.quantity for scan in scans if scan.status != "active"),
        "affected_parts": sorted(affected.values(), key=lambda item: str(item["part_key"])),
        "current_status": {key: value for key, value in status.items() if key != "rows"},
        "preview_status": preview_status,
    }


def outbound_ops_health(session: Session) -> dict[str, object]:
    settings = get_settings()
    is_sqlite = session.get_bind().dialect.name == "sqlite"
    if is_sqlite:
        db_quick_check_row = session.exec(text("PRAGMA quick_check")).one()
        if isinstance(db_quick_check_row, tuple):
            db_quick_check_value = db_quick_check_row[0]
        elif hasattr(db_quick_check_row, "_mapping"):
            db_quick_check_value = next(iter(db_quick_check_row._mapping.values()))
        else:
            db_quick_check_value = db_quick_check_row
        db_quick_check = str(db_quick_check_value)
    else:
        # PostgreSQL does not support PRAGMA; validate connectivity with a cheap heartbeat.
        heartbeat = session.exec(text("SELECT 1")).one()
        if isinstance(heartbeat, tuple):
            heartbeat_value = heartbeat[0]
        elif hasattr(heartbeat, "_mapping"):
            heartbeat_value = next(iter(heartbeat._mapping.values()))
        else:
            heartbeat_value = heartbeat
        db_quick_check = "ok" if int(heartbeat_value) == 1 else "degraded"

    if is_sqlite:
        database_file = _public_path(settings.database_path)
        backup_dir = settings.database_path.parent / "backups"
        backup_globs = ["*.db"]
        database_kind = "sqlite"
    else:
        database_file = "postgresql://***"
        backup_dir = Path(__file__).resolve().parents[3] / "data" / "backups"
        backup_globs = ["*.dump", "*.backup"]
        database_kind = "postgresql"
    backups: list[Path] = []
    for pattern in backup_globs:
        backups.extend(backup_dir.glob(pattern))
    backups = sorted(backups, key=lambda path: path.stat().st_mtime, reverse=True)
    latest_backup = backups[0] if backups else None
    scan_rows = session.exec(
        select(
            OutboundScan.status, func.count(OutboundScan.id), func.sum(OutboundScan.quantity)
        ).group_by(OutboundScan.status)
    ).all()
    snapshot_count = int(session.exec(select(func.count(OutboundProgressSnapshot.id))).one() or 0)
    error_records = session.exec(
        select(Record)
        .where(Record.last_error.is_not(None))
        .order_by(Record.updated_at.desc(), Record.id.desc())
        .limit(5)
    ).all()
    try:
        outbound_overview = outbound_orders_overview(session)
        default_order = outbound_order_status("SO202604210135", session)
    except RuntimeError as exc:
        outbound_overview = {"error": str(exc)}
        default_order = {"error": str(exc)}
    return {
        "status": "ok" if db_quick_check == "ok" else "degraded",
        "database": {
            "kind": database_kind,
            "file": database_file,
            "quick_check": db_quick_check,
        },
        "runtime": {
            "ocr_provider": settings.ocr_provider,
            "enable_barcode": settings.enable_barcode,
            "enable_saas_submit": settings.enable_saas_submit,
            "dry_run": settings.dry_run,
            "host_name": socket.gethostname().split(".")[0],
        },
        "backups": {
            "count": len(backups),
            "latest": None
            if latest_backup is None
            else {
                "file": _public_path(latest_backup),
                "size_bytes": latest_backup.stat().st_size,
                "modified_at": datetime.fromtimestamp(
                    latest_backup.stat().st_mtime, UTC
                ).isoformat(),
            },
        },
        "outbound": {
            "totals": outbound_overview.get("totals", {})
            if isinstance(outbound_overview, dict)
            else {},
            "default_order": {key: value for key, value in default_order.items() if key != "rows"}
            if isinstance(default_order, dict)
            else {},
            "scan_status": {
                str(status): {"count": int(count or 0), "quantity": int(quantity or 0)}
                for status, count, quantity in scan_rows
            },
            "snapshot_count": snapshot_count,
        },
        "recent_errors": [
            {
                "id": record.id,
                "status": record.status,
                "has_error": bool(record.last_error),
                "updated_at": record.updated_at.isoformat() if record.updated_at else None,
            }
            for record in error_records
        ],
    }


def void_outbound_batch(
    *,
    order_no: str,
    batch_id: str,
    operator_id: str,
    session: Session,
) -> dict[str, object]:
    selected_order = normalize_order_no(order_no)
    batch_id = batch_id.strip()
    if not batch_id:
        raise RuntimeError("batch_id is required")
    scans = session.exec(
        select(OutboundScan).where(
            OutboundScan.order_no == selected_order,
            OutboundScan.batch_id == batch_id,
            OutboundScan.status == "active",
        )
    ).all()
    if not scans:
        raise RuntimeError(f"active outbound batch not found: {selected_order} {batch_id}")
    now = datetime.now(UTC)
    for scan in scans:
        if scan.location_code:
            try:
                inbound_inventory(
                    part_key=scan.part_code,
                    location_code=scan.location_code,
                    quantity=scan.quantity,
                    operator_id=operator_id,
                    reason="void_outbound_batch_revert",
                    session=session,
                )
            except RuntimeError:
                pass
        scan.status = "voided"
        scan.void_reason = "batch_void"
        scan.voided_by = (operator_id.strip() or "self")[:80]
        scan.voided_at = now
        session.add(scan)
    session.commit()
    status = outbound_order_status(selected_order, session)
    snapshot = save_outbound_progress_snapshot(
        order_no=selected_order,
        event="batch_voided",
        operator_id=operator_id,
        session=session,
        status=status,
        batch_id=batch_id,
        detail={
            "batch_id": batch_id,
            "voided_scan_ids": [scan.id for scan in scans],
            "voided_quantity": sum(scan.quantity for scan in scans),
        },
    )
    return {
        "voided": True,
        "batch_id": batch_id,
        "voided_scan_count": len(scans),
        "voided_quantity": sum(scan.quantity for scan in scans),
        "snapshot": _snapshot_to_payload(snapshot),
        "order_status": status,
    }


def void_outbound_scan(
    *,
    scan_id: int,
    operator_id: str,
    reason: str,
    session: Session,
) -> dict[str, object]:
    scan = session.get(OutboundScan, scan_id)
    if scan is None:
        raise RuntimeError(f"outbound scan not found: {scan_id}")
    if scan.status != "voided":
        scan.status = "voided"
        scan.void_reason = (reason.strip() or "operator_void")[:200]
        scan.voided_by = (operator_id.strip() or "self")[:80]
        scan.voided_at = datetime.now(UTC)
        session.add(scan)
        session.commit()
        session.refresh(scan)
        if scan.location_code:
            try:
                inbound_inventory(
                    part_key=scan.part_code,
                    location_code=scan.location_code,
                    quantity=scan.quantity,
                    operator_id=operator_id,
                    reason="void_outbound_scan_revert",
                    session=session,
                )
            except RuntimeError:
                pass
    status = outbound_order_status(scan.order_no, session)
    snapshot = save_outbound_progress_snapshot(
        order_no=scan.order_no,
        event="scan_voided",
        operator_id=operator_id,
        session=session,
        status=status,
        scan_id=scan.id,
        detail={"part_key": scan.part_code, "quantity": scan.quantity, "reason": reason},
    )
    return {
        "voided": True,
        "scan": _scan_to_payload(scan),
        "snapshot": _snapshot_to_payload(snapshot),
        "order_status": status,
    }


def set_outbound_part_quantity(
    *,
    order_no: str,
    part_key: str,
    quantity: int,
    operator_id: str,
    session: Session,
    reason: str = "manual_set",
    batch_id: str | None = None,
) -> dict[str, object]:
    selected_order = normalize_order_no(order_no)
    normalized_part = compact_part_code(part_key)
    required_rows = _order_required_rows(selected_order)
    matched_row = required_rows.get(normalized_part)
    if matched_row is None:
        raise RuntimeError(f"part not found in order: {selected_order} {part_key}")
    if matched_row["unknown_quantity"]:
        raise RuntimeError(f"part quantity unreadable: {selected_order} {part_key}")
    required_qty = int(matched_row["required_qty"])
    target_qty = max(0, min(int(quantity), required_qty))
    active_scans = session.exec(
        select(OutboundScan).where(
            OutboundScan.order_no == selected_order,
            OutboundScan.part_code == normalized_part,
            OutboundScan.status == "active",
        )
    ).all()
    if batch_id:
        existing = [scan for scan in active_scans if _is_same_batch_scope(scan.batch_id, batch_id)]
        preserved_qty = sum(
            scan.quantity
            for scan in active_scans
            if not _is_same_batch_scope(scan.batch_id, batch_id)
        )
    else:
        existing = active_scans
        preserved_qty = 0
    now = datetime.now(UTC)
    for scan in existing:
        if scan.location_code:
            try:
                inbound_inventory(
                    part_key=normalized_part,
                    location_code=scan.location_code,
                    quantity=scan.quantity,
                    operator_id=operator_id,
                    reason=f"{reason}_void_revert",
                    session=session,
                )
            except RuntimeError:
                pass
        scan.status = "voided"
        scan.void_reason = reason[:200]
        scan.voided_by = (operator_id.strip() or "self")[:80]
        scan.voided_at = now
        session.add(scan)
    scan = None
    new_qty = max(0, target_qty - preserved_qty)
    if new_qty > 0:
        selected_location, available_locations, location_matches = _select_location_for_outbound(
            location_code=None,
            required_row=matched_row,
        )
        if available_locations and not selected_location:
            selected_location = available_locations[0]
            location_matches = True
        if available_locations and not location_matches:
            raise RuntimeError(
                f"invalid location selection for manual quantity update: {normalized_part}"
            )
        if selected_location:
            _bootstrap_inventory_if_missing(
                session,
                part_key=normalized_part,
                location_code=selected_location,
                operator_id=operator_id,
                reason="legacy_bootstrap_from_manual_quantity",
                seed_quantity=max(new_qty, required_qty),
            )
            outbound_inventory(
                part_key=normalized_part,
                location_code=selected_location,
                quantity=new_qty,
                operator_id=operator_id,
                reason=reason,
                session=session,
                order_no=selected_order,
            )
        scan = OutboundScan(
            order_no=selected_order,
            part_code=normalized_part,
            location_code=selected_location,
            source_code=_manual_source_code(normalized_part),
            matched_code=str(matched_row["part_code"]),
            quantity=new_qty,
            status="active",
            operator_id=(operator_id.strip() or "self")[:80],
            batch_id=(batch_id or "")[:120] or None,
            record_id=None,
        )
        session.add(scan)
    session.commit()
    if scan is not None:
        session.refresh(scan)
    status = outbound_order_status(selected_order, session)
    snapshot = save_outbound_progress_snapshot(
        order_no=selected_order,
        event=reason,
        operator_id=operator_id,
        session=session,
        status=status,
        scan_id=scan.id if scan is not None else None,
        batch_id=batch_id,
        detail={
            "part_key": normalized_part,
            "target_quantity": target_qty,
            "new_quantity": new_qty,
            "preserved_quantity": preserved_qty,
            "batch_id": batch_id,
        },
    )
    return {
        "updated": True,
        "target_quantity": target_qty,
        "scan": _scan_to_payload(scan) if scan is not None else None,
        "snapshot": _snapshot_to_payload(snapshot),
        "matched_part": matched_row,
        "order_status": status,
    }


def complete_outbound_order(
    *,
    order_no: str,
    operator_id: str,
    session: Session,
) -> dict[str, object]:
    selected_order = normalize_order_no(order_no)
    required_rows = _order_required_rows(selected_order)
    if not required_rows:
        raise RuntimeError(f"order not found: {selected_order}")
    batch_id = f"complete-{selected_order}-{uuid.uuid4().hex[:12]}"
    for part_key, row in required_rows.items():
        if row["unknown_quantity"]:
            continue
        set_outbound_part_quantity(
            order_no=selected_order,
            part_key=part_key,
            quantity=int(row["required_qty"]),
            operator_id=operator_id,
            session=session,
            reason="complete_order",
            batch_id=batch_id,
        )
    status = outbound_order_status(selected_order, session)
    completed_at = datetime.now(UTC)
    snapshot = save_outbound_progress_snapshot(
        order_no=selected_order,
        event="complete_order_finished",
        operator_id=operator_id,
        session=session,
        status=status,
        batch_id=batch_id,
        completed_at=completed_at,
        detail={"line_total": len(required_rows), "batch_id": batch_id},
    )
    return {
        "completed": True,
        "batch_id": batch_id,
        "snapshot": _snapshot_to_payload(snapshot),
        "order_status": status,
    }


def rollback_outbound_order(
    *,
    order_no: str,
    operator: User,
    session: Session,
) -> dict[str, object]:
    selected_order = normalize_order_no(order_no)
    latest_completion = session.exec(
        select(OutboundProgressSnapshot)
        .where(
            OutboundProgressSnapshot.order_no == selected_order,
            OutboundProgressSnapshot.event == "complete_order_finished",
        )
        .order_by(OutboundProgressSnapshot.created_at.desc(), OutboundProgressSnapshot.id.desc())
    ).first()
    if latest_completion is None:
        raise RuntimeError(f"no completion snapshot found for order: {selected_order}")

    completed_at = latest_completion.completed_at or latest_completion.created_at
    if completed_at is None:
        raise RuntimeError(f"invalid completion snapshot for order: {selected_order}")
    if completed_at.tzinfo is None:
        completed_at = completed_at.replace(tzinfo=UTC)
    else:
        completed_at = completed_at.astimezone(UTC)

    now = datetime.now(UTC)
    rollback_window_minutes = max(1, int(get_settings().rollback_window_minutes or 30))
    rollback_deadline = completed_at + timedelta(minutes=rollback_window_minutes)
    if now > rollback_deadline:
        raise RuntimeError(
            f"rollback window expired for order {selected_order}: "
            f"completed_at={completed_at.isoformat()} deadline={rollback_deadline.isoformat()}"
        )

    active_scans = session.exec(
        select(OutboundScan)
        .where(OutboundScan.order_no == selected_order, OutboundScan.status == "active")
        .order_by(OutboundScan.id.asc())
    ).all()
    if not active_scans:
        raise RuntimeError(f"no active scans to rollback for order: {selected_order}")

    now = datetime.now(UTC)
    restored_count = 0
    restored_quantity = 0
    movement_rows: list[InventoryMovement] = []
    for scan in active_scans:
        if scan.location_code:
            normalized_location = _normalize_location_code(scan.location_code)
            location = _get_or_create_inventory_location(
                session,
                part_key=scan.part_code,
                location_code=normalized_location,
            )
            before_qty = int(location.quantity)
            after_qty = before_qty + int(scan.quantity)
            location.quantity = after_qty
            location.zero_stock = after_qty == 0
            if location.zero_stock:
                location.status = "zero_stock"
            elif location.status in {"zero_stock", "retired"}:
                location.status = "active"
            location.updated_at = now
            session.add(location)
            session.flush()

            movement = InventoryMovement(
                movement_type="inbound",
                part_key=scan.part_code,
                location_code=normalized_location,
                order_no=selected_order,
                scan_id=scan.id,
                quantity_delta=int(scan.quantity),
                before_qty=before_qty,
                after_qty=after_qty,
                operator_id=(operator.username or "self")[:80],
                reason="rollback_outbound_order_revert",
            )
            session.add(movement)
            movement_rows.append(movement)
            restored_count += 1
            restored_quantity += int(scan.quantity)

        scan.status = "voided"
        scan.void_reason = "rollback"
        scan.voided_by = (operator.username or "self")[:80]
        scan.voided_at = now
        session.add(scan)
    session.flush()

    status = outbound_order_status(selected_order, session)
    active_count, active_quantity = _active_scan_summary(session, selected_order)
    rollback_snapshot = OutboundProgressSnapshot(
        order_no=selected_order,
        event="rollback_completed",
        required_total=int(status.get("required_total") or 0),
        scanned_total=int(status.get("scanned_total") or 0),
        remaining_total=int(status.get("remaining_total") or 0),
        line_total=int(status.get("line_total") or 0),
        complete_line_total=int(status.get("complete_line_total") or 0),
        active_scan_count=active_count,
        active_scan_quantity=active_quantity,
        operator_id=(operator.username or "self")[:80],
        batch_id=latest_completion.batch_id,
        scan_id=None,
        detail_json=json.dumps(
            {
                "rollback_source_snapshot_id": latest_completion.id,
                "rollback_window_minutes": rollback_window_minutes,
                "voided_scan_ids": [scan.id for scan in active_scans],
                "voided_quantity": sum(int(scan.quantity) for scan in active_scans),
                "restored_inventory_rows": restored_count,
                "restored_inventory_quantity": restored_quantity,
            },
            ensure_ascii=False,
            default=str,
        ),
    )
    session.add(rollback_snapshot)
    session.flush()

    audit = AuditLog(
        event_type="outbound_rollback",
        actor_user_id=operator.id,
        actor_username=operator.username,
        target_type="outbound_order",
        target_id=selected_order,
        action="outbound.rollback",
        reason="supervisor rollback",
        success=True,
        detail_json=json.dumps(
            {
                "order_no": selected_order,
                "rollback_source_snapshot_id": latest_completion.id,
                "voided_scan_count": len(active_scans),
                "voided_quantity": sum(int(scan.quantity) for scan in active_scans),
                "rollback_window_minutes": rollback_window_minutes,
                "completed_at": completed_at.isoformat(),
                "rolled_back_at": now.isoformat(),
            },
            ensure_ascii=False,
            default=str,
        ),
    )
    session.add(audit)
    session.commit()
    session.refresh(rollback_snapshot)
    session.refresh(audit)

    return {
        "rolled_back": True,
        "order_no": selected_order,
        "voided_scan_count": len(active_scans),
        "voided_quantity": sum(int(scan.quantity) for scan in active_scans),
        "restored_inventory_rows": restored_count,
        "restored_inventory_quantity": restored_quantity,
        "rollback_window_minutes": rollback_window_minutes,
        "completed_at": completed_at.isoformat(),
        "rolled_back_at": now.isoformat(),
        "snapshot": _snapshot_to_payload(rollback_snapshot),
        "audit_log_id": audit.id,
        "order_status": status,
    }


def sync_outbound_completion_marks(
    *,
    text: str,
    operator_id: str,
    session: Session,
    order_no: str | None = None,
) -> dict[str, object]:
    marks = parse_outbound_completion_marks(text)
    selected_order = normalize_order_no(order_no) if order_no else ""
    applied = []
    skipped = []
    for mark in marks:
        if selected_order and mark.order_no != selected_order:
            skipped.append({**mark.__dict__, "reason": "outside_selected_order"})
            continue
        part_key = compact_part_code(mark.part_code)
        required_rows = _order_required_rows(mark.order_no)
        row = required_rows.get(part_key)
        if row is None:
            skipped.append({**mark.__dict__, "reason": "part_not_found"})
            continue
        if row["unknown_quantity"]:
            skipped.append({**mark.__dict__, "reason": "quantity_unreadable"})
            continue
        if mark.quantity is not None and mark.quantity != int(row["required_qty"]):
            skipped.append(
                {
                    **mark.__dict__,
                    "reason": "quantity_mismatch",
                    "required_qty": int(row["required_qty"]),
                }
            )
            continue
        result = set_outbound_part_quantity(
            order_no=mark.order_no,
            part_key=part_key,
            quantity=int(row["required_qty"]),
            operator_id=operator_id,
            session=session,
            reason="completion_mark_sync",
            batch_id=f"marks-{selected_order or mark.order_no}",
        )
        applied.append(
            {
                **mark.__dict__,
                "part_key": part_key,
                "target_quantity": result["target_quantity"],
            }
        )
    touched_orders = sorted({mark["order_no"] for mark in applied})
    return {
        "parsed_count": len(marks),
        "applied_count": len(applied),
        "skipped_count": len(skipped),
        "applied": applied,
        "skipped": skipped,
        "orders": [outbound_order_status(order_no, session) for order_no in touched_orders],
    }
