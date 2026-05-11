from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from zipfile import BadZipFile

from backend.app.config import get_settings


@dataclass(frozen=True)
class MaterialMatch:
    ruiyun_part_number: str
    sku: str
    matched_input: str
    matched_field: str
    material_code: str = ""
    byd_part_number: str = ""
    description: str = ""
    material_type: str = ""
    product_name: str = ""
    carton_size: str = ""


def normalize_material_code(value: str | None) -> str:
    if value is None:
        return ""
    return "".join(ch for ch in value.upper() if ch.isalnum())


def _cell_text(value: object) -> str:
    return str(value).strip() if value is not None else ""


def _normalize_header(value: str) -> str:
    return "".join(ch for ch in value.strip().lower() if ch.isalnum() or "\u4e00" <= ch <= "\u9fff")


def _find_header_indexes(headers: list[str]) -> dict[str, int] | None:
    normalized = [_normalize_header(header) for header in headers]
    candidates = {
        "ruiyun_part_number": {"瑞云料号", "ruiyun", "ruiyun料号", "料号"},
        "sku": {"sku", "松灵物料", "松灵料号", "松灵sku", "物料", "物料号"},
        "material_code": {"物料编码", "物料代码", "物料编码partcode", "partcode", "materialcode"},
        "byd_part_number": {"byd料号", "byd物料", "byd物料号", "bydpartnumber", "byd"},
        "description": {"描述", "物料描述", "说明", "description"},
        "material_type": {"类型", "物料类型", "type", "materialtype"},
        "product_name": {"品名", "产品名称", "名称", "name", "productname"},
        "carton_size": {"箱规尺寸", "箱规", "尺寸", "包装尺寸", "cartonsize", "boxsize"},
    }
    indexes = {
        field: index
        for field, names in candidates.items()
        if (index := next((i for i, header in enumerate(normalized) if header in names), None))
        is not None
    }
    ruiyun_index = indexes.get("ruiyun_part_number")
    sku_index = indexes.get("sku")
    if ruiyun_index is None or sku_index is None or ruiyun_index == sku_index:
        return None
    return indexes


def load_material_matches(path: Path) -> list[MaterialMatch]:
    if not path.exists():
        return []
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        return []
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (BadZipFile, OSError, ValueError):
        return []

    matches: list[MaterialMatch] = []
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                continue
            header_indexes = _find_header_indexes([_cell_text(cell) for cell in header])
            if header_indexes is None:
                continue
            ruiyun_index = header_indexes["ruiyun_part_number"]
            sku_index = header_indexes["sku"]
            for row in rows:
                values = list(row)
                ruiyun = _cell_text(values[ruiyun_index]) if ruiyun_index < len(values) else ""
                sku = _cell_text(values[sku_index]) if sku_index < len(values) else ""
                if ruiyun and sku:
                    extra = {
                        field: _cell_text(values[index]) if index < len(values) else ""
                        for field, index in header_indexes.items()
                        if field not in {"ruiyun_part_number", "sku"}
                    }
                    matches.append(
                        MaterialMatch(
                            ruiyun_part_number=ruiyun,
                            sku=sku,
                            matched_input="",
                            matched_field="",
                            **extra,
                        )
                    )
    finally:
        workbook.close()
    return matches


@lru_cache(maxsize=4)
def _cached_material_matches(path_text: str, mtime_ns: int, size: int) -> tuple[MaterialMatch, ...]:
    return tuple(load_material_matches(Path(path_text)))


def material_catalog() -> list[MaterialMatch]:
    path = get_settings().material_mapping_file
    try:
        stat = path.stat()
    except OSError:
        return []
    return list(_cached_material_matches(str(path), stat.st_mtime_ns, stat.st_size))


def find_material_matches(text: str, limit: int = 5) -> list[MaterialMatch]:
    compact_text = normalize_material_code(text)
    if not compact_text:
        return []

    found: list[MaterialMatch] = []
    seen: set[tuple[str, str]] = set()
    for item in material_catalog():
        candidates = [
            (item.ruiyun_part_number, "ruiyun_part_number"),
            (item.sku, "sku"),
            (item.material_code, "material_code"),
            (item.byd_part_number, "byd_part_number"),
        ]
        for raw_value, field in candidates:
            normalized = normalize_material_code(raw_value)
            if normalized and normalized in compact_text:
                key = (item.ruiyun_part_number, item.sku)
                if key not in seen:
                    found.append(
                        MaterialMatch(
                            ruiyun_part_number=item.ruiyun_part_number,
                            sku=item.sku,
                            matched_input=raw_value,
                            matched_field=field,
                        )
                    )
                    seen.add(key)
                break
        if len(found) >= limit:
            break
    return found


def search_material_catalog(query: str = "", limit: int | None = None) -> list[MaterialMatch]:
    compact_query = normalize_material_code(query)
    text_query = query.strip().lower()
    rows = material_catalog()
    if compact_query or text_query:
        filtered = []
        for item in rows:
            code_values = (
                item.material_code,
                item.sku,
                item.byd_part_number,
                item.ruiyun_part_number,
            )
            text_values = (
                item.description,
                item.material_type,
                item.product_name,
                item.carton_size,
            )
            code_hit = any(compact_query in normalize_material_code(value) for value in code_values)
            text_hit = any(text_query in value.lower() for value in text_values if value)
            if code_hit or text_hit:
                filtered.append(item)
        rows = filtered
    if limit is not None:
        return rows[:limit]
    return rows


def material_match_to_dict(item: MaterialMatch) -> dict[str, str]:
    return {
        "material_code": item.material_code,
        "sku": item.sku,
        "byd_part_number": item.byd_part_number,
        "ruiyun_part_number": item.ruiyun_part_number,
        "description": item.description,
        "material_type": item.material_type,
        "product_name": item.product_name,
        "carton_size": item.carton_size,
        "matched_input": item.matched_input,
        "matched_field": item.matched_field,
    }


def material_matches_to_text(matches: list[MaterialMatch]) -> str:
    if not matches:
        return ""
    lines = ["MATERIAL MATCHES:"]
    for item in matches:
        lines.append(f"RUIYUN: {item.ruiyun_part_number}")
        lines.append(f"SKU: {item.sku}")
    return "\n".join(lines)


def clear_material_mapping_cache() -> None:
    _cached_material_matches.cache_clear()
