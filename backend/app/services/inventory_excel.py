from __future__ import annotations

import csv
from io import BytesIO, StringIO
from pathlib import Path

REQUIRED_COLUMNS = {"part_key", "location_code", "quantity"}
OPTIONAL_COLUMNS = {"factory_id"}
SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _normalize_cell(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def _quantity_value(value: object) -> int:
    if value is None:
        raise RuntimeError("quantity is required")
    text = str(value).strip()
    if not text:
        raise RuntimeError("quantity is required")
    try:
        number = float(text)
    except ValueError as exc:
        raise RuntimeError(f"invalid quantity: {value}") from exc
    if not number.is_integer():
        raise RuntimeError("quantity must be an integer")
    quantity = int(number)
    if quantity < 0:
        raise RuntimeError("quantity must be >= 0")
    return quantity


def _row_payload(raw: dict[str, object]) -> dict[str, object] | None:
    if not any(str(value or "").strip() for value in raw.values()):
        return None
    missing = [column for column in sorted(REQUIRED_COLUMNS) if column not in raw]
    if missing:
        raise RuntimeError(f"missing required columns: {', '.join(missing)}")

    row: dict[str, object] = {
        "part_key": str(raw["part_key"] or "").strip(),
        "location_code": str(raw["location_code"] or "").strip(),
        "quantity": _quantity_value(raw["quantity"]),
    }
    factory_id = str(raw.get("factory_id") or "").strip()
    if factory_id:
        row["factory_id"] = factory_id
    return row


def _rows_from_table(headers: list[object], rows: list[list[object]]) -> list[dict[str, object]]:
    normalized_headers = [_normalize_header(header) for header in headers]
    if not any(normalized_headers):
        raise RuntimeError("header row is required")
    missing = [column for column in sorted(REQUIRED_COLUMNS) if column not in normalized_headers]
    if missing:
        raise RuntimeError(f"missing required columns: {', '.join(missing)}")

    parsed_rows: list[dict[str, object]] = []
    allowed_columns = REQUIRED_COLUMNS | OPTIONAL_COLUMNS
    for values in rows:
        raw = {
            header: _normalize_cell(values[index]) if index < len(values) else ""
            for index, header in enumerate(normalized_headers)
            if header in allowed_columns
        }
        payload = _row_payload(raw)
        if payload is not None:
            parsed_rows.append(payload)
    return parsed_rows


def _parse_csv(content: bytes) -> list[dict[str, object]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    if reader.fieldnames is None:
        raise RuntimeError("header row is required")
    headers = [_normalize_header(header) for header in reader.fieldnames]
    rows = [[row.get(header, "") for header in reader.fieldnames] for row in reader]
    return _rows_from_table(headers, rows)


def _parse_xlsx(content: bytes) -> list[dict[str, object]]:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        headers = list(next(iterator))
    except StopIteration as exc:
        raise RuntimeError("header row is required") from exc
    rows = [list(row) for row in iterator]
    return _rows_from_table(headers, rows)


def parse_inventory_file_rows(*, filename: str, content: bytes) -> list[dict[str, object]]:
    extension = Path(filename or "").suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise RuntimeError("unsupported inventory file type")
    if extension == ".csv":
        return _parse_csv(content)
    return _parse_xlsx(content)
